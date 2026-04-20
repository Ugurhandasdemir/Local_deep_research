"""
Vektor Veritabani Benchmark Scripti
------------------------------------
Amac: 4 adet local embedding modelini (bge_squad, e5_base, e5_small, mpnet_multi)
      5 farkli vektor veritabaninda (Milvus, Qdrant, Chroma, LanceDB, Weaviate)
      test etmek.

Sistem: 4GB VRAM. Bu yuzden modeller tek tek GPU'ya alinir,
        is bitince bellekten silinir.

Kullanim:
    python write_vector_database.py --mode test    # 20 belge ile hizli test
    python write_vector_database.py --mode full    # butun belgelerle tam benchmark
"""

import os
import gc
import json
import time
import shutil
import argparse
import warnings
from datetime import datetime

import numpy as np
import torch
from sentence_transformers import SentenceTransformer

# Veritabani kutuphaneleri
import lancedb
import chromadb
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct, SearchParams
from pymilvus import MilvusClient
import weaviate
from weaviate.classes.config import Property, DataType, Configure

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings("ignore")


# ---------------------------------------------------------------
# Ayarlar
# ---------------------------------------------------------------
BASE_DIR = "/home/ugo/Documents/Python/bitirememe projesi"
MODELS_DIR = os.path.join(BASE_DIR, "models")
DB_DIR = os.path.join(BASE_DIR, "DB")
DATA_FILE = os.path.join(BASE_DIR, "CUSTOM_DATASET", "metin_dosyasi.json")

# Tum embedding modelleri — hepsi 4GB VRAM'e rahatlikla sigar.
# "path" HuggingFace ismi (online indirir) veya local klasor olabilir; ikisi de calisir.
MODELS = {
    # --- LOCAL fine-tuned modeller ---
    "e5_small":           {"path": os.path.join(MODELS_DIR, "e5_small", "model"),    "dim": 384},
    "mpnet_multi":        {"path": os.path.join(MODELS_DIR, "mpnet_multi", "model"), "dim": 768},
    "e5_base":            {"path": os.path.join(MODELS_DIR, "e5_base", "model"),     "dim": 768},
    "bge_squad":          {"path": os.path.join(MODELS_DIR, "bge_squad_model"),      "dim": 1024},
    "qwen_lora":          {"path": os.path.join(MODELS_DIR, "qwen_lora"),             "dim": 1024},
    "snowflake_arctic_l": {"path": os.path.join(MODELS_DIR, "snowflake-arctic-embed-l-v2.0"), "dim": 1024},
    "all_mini_l6":        {"path": os.path.join(MODELS_DIR, "all_mini_l6_v2"),         "dim": 384},

    # --- Fine-tuned modellerin EGITILMEMIS (base) HuggingFace versiyonlari ---
    # Karsilastirma: bu modeller yukarıdaki local fine-tuned surumlerin baz modelleridir.
    "e5_small_base":          {"path": "intfloat/multilingual-e5-small",                   "dim": 384},   # e5_small icin base
    "mpnet_multi_base":       {"path": "sentence-transformers/paraphrase-multilingual-mpnet-base-v2", "dim": 768},  # mpnet_multi icin base
    "e5_base_base":           {"path": "intfloat/multilingual-e5-base",                    "dim": 768},   # e5_base icin base
    "bge_squad_base":         {"path": "BAAI/bge-large-en-v1.5",                           "dim": 1024},  # bge_squad icin base
    "qwen_lora_base":         {"path": "Qwen/Qwen3-Embedding-0.6B",                        "dim": 1024},  # qwen_lora icin base
    "snowflake_arctic_l_base": {"path": "Snowflake/snowflake-arctic-embed-l-v2.0",         "dim": 1024},  # snowflake_arctic_l icin base
    "all_mini_l6_base":       {"path": "sentence-transformers/all-MiniLM-L6-v2",           "dim": 384},   # all_mini_l6 icin base

    # --- Diger HuggingFace hazir modeller (ilk calistirmada indirilir, sonra cache'lenir) ---
    "minilm_l12":         {"path": "sentence-transformers/all-MiniLM-L12-v2",                     "dim": 384},
    "mpnet_base":         {"path": "sentence-transformers/all-mpnet-base-v2",                     "dim": 768},
    "distilroberta":      {"path": "sentence-transformers/all-distilroberta-v1",                  "dim": 768},
    "multi_qa_minilm":    {"path": "sentence-transformers/multi-qa-MiniLM-L6-cos-v1",             "dim": 384},
    "multi_qa_mpnet":     {"path": "sentence-transformers/multi-qa-mpnet-base-dot-v1",            "dim": 768},
    "paraphrase_multi":   {"path": "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2", "dim": 384},
    "bge_small_en":       {"path": "BAAI/bge-small-en-v1.5",                                      "dim": 384},
    "bge_base_en":        {"path": "BAAI/bge-base-en-v1.5",                                       "dim": 768},
    "gte_small":          {"path": "thenlper/gte-small",                                          "dim": 384},
    "gte_base":           {"path": "thenlper/gte-base",                                           "dim": 768},
    "e5_small_hf":        {"path": "intfloat/e5-small-v2",                                        "dim": 384},
    "e5_base_hf":         {"path": "intfloat/e5-base-v2",                                         "dim": 768},
}

TEST_QUERIES = [
    "artificial intelligence healthcare applications",
    "machine learning medical diagnosis systems",
    "deep learning neural network architectures",
    "natural language processing techniques",
    "computer vision medical imaging analysis",
    "reinforcement learning robotics control",
    "transformer models attention mechanism",
    "convolutional neural networks image classification",
    "recurrent neural networks sequence modeling",
    "generative adversarial networks image synthesis",
]

# Arama benchmark ayarlari — tek yerden duzenlenebilsin diye
WARMUP_RUNS = 2
TEST_RUNS = 10     # Her arama icin 10 kez calistir, ortalama al
BATCH_SIZE = 32    # 4GB VRAM icin orta boy batch

OUTPUT_JSON = os.path.join(BASE_DIR, "multi_model_benchmark_results.json")
OUTPUT_XLSX = os.path.join(BASE_DIR, "multi_model_benchmark_results.xlsx")


# ---------------------------------------------------------------
# Yardimci fonksiyonlar
# ---------------------------------------------------------------
def free_memory():
    """GPU + RAM bellegini temizle"""
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()


def print_header(title):
    print("\n" + "=" * 70)
    print(" " + title)
    print("=" * 70)


def load_documents(limit=None):
    """CUSTOM_DATASET'teki pdf metinlerini yukle, 500 karakterlik chunklara bol."""
    print_header("BELGELER YUKLENIYOR")

    if not os.path.exists(DATA_FILE):
        raise FileNotFoundError(f"Veri dosyasi bulunamadi: {DATA_FILE}")

    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    pdf_docs = data.get("documents", [])
    chunks = []
    chunk_size = 500
    overlap = 50
    step = chunk_size - overlap

    for doc in pdf_docs:
        text = doc.get("full_text", "").strip()
        filename = doc.get("filename", "unknown")
        if not text:
            continue
        for i in range(0, len(text), step):
            chunk = text[i:i + chunk_size].strip()
            if len(chunk) > 50:
                chunks.append({"text": chunk, "source": filename})

    if limit is not None:
        chunks = chunks[:limit]
        print(f"  TEST modu: ilk {limit} chunk kullanilacak")

    print(f"  {len(pdf_docs)} PDF -> {len(chunks)} chunk")
    return chunks


def measure_time(func):
    """Bir arama fonksiyonunu warmup + test kez calistirip istatistik dondur."""
    # Once warmup — cache/JIT isinsin
    for _ in range(WARMUP_RUNS):
        try:
            func()
        except Exception as e:
            return {"error": str(e)}

    times = []
    for _ in range(TEST_RUNS):
        t0 = time.time()
        try:
            func()
        except Exception as e:
            return {"error": str(e)}
        times.append(time.time() - t0)

    return {
        "avg_time": float(np.mean(times)),
        "min_time": float(np.min(times)),
        "max_time": float(np.max(times)),
        "std_time": float(np.std(times)),
        "p50_time": float(np.percentile(times, 50)),
        "p95_time": float(np.percentile(times, 95)),
    }


# ---------------------------------------------------------------
# Embedding hesaplama (tek model, 4GB VRAM icin)
# ---------------------------------------------------------------
def encode_model(model_path, texts, queries):
    """
    Bir modeli yukle, belgeleri ve sorgulari encode et, modeli sil.
    4GB VRAM'e tek bir model sigar, bu yuzden her model sonrasi bellekten atilir.
    """
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"  Model yukleniyor: {model_path}  (device={device})")

    model = SentenceTransformer(model_path, device=device)

    print(f"  {len(texts)} belge encode ediliyor...")
    doc_embs = model.encode(
        texts,
        show_progress_bar=True,
        batch_size=BATCH_SIZE,
        convert_to_numpy=True,
        normalize_embeddings=True,
    )

    query_embs = model.encode(
        queries,
        show_progress_bar=False,
        convert_to_numpy=True,
        normalize_embeddings=True,
    )

    # Bellegi hemen bosalt — siradaki model icin
    del model
    free_memory()

    return doc_embs.tolist(), query_embs.tolist()


# ---------------------------------------------------------------
# VERITABANINA YAZMA — her fonksiyon 1 modelin embeddinglerini yazar
# ---------------------------------------------------------------
def write_to_milvus(model_key, docs, embeddings):
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "milvus", f"{model_key}_db.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)

    if os.path.exists(db_path):
        os.remove(db_path)

    client = MilvusClient(db_path)
    t0 = time.time()
    client.create_collection(
        collection_name=collection,
        dimension=len(embeddings[0]),
        metric_type="COSINE",
    )
    data = [
        {"id": i, "vector": emb, "text": d["text"][:500], "source": d["source"]}
        for i, (d, emb) in enumerate(zip(docs, embeddings))
    ]
    for i in range(0, len(data), 100):
        client.insert(collection_name=collection, data=data[i:i + 100])
    return time.time() - t0


def write_to_qdrant(model_key, docs, embeddings):
    client = QdrantClient(host="localhost", port=6333, timeout=120)
    collection = f"docs_{model_key}"

    try:
        client.delete_collection(collection)
    except Exception:
        pass

    t0 = time.time()
    client.create_collection(
        collection_name=collection,
        vectors_config=VectorParams(size=len(embeddings[0]), distance=Distance.COSINE),
    )
    points = [
        PointStruct(id=i, vector=emb, payload={"text": d["text"], "source": d["source"]})
        for i, (d, emb) in enumerate(zip(docs, embeddings))
    ]
    for i in range(0, len(points), 100):
        client.upsert(collection_name=collection, points=points[i:i + 100])
    return time.time() - t0


def write_to_chromadb(model_key, docs, embeddings):
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "chromadb", f"{model_key}_db")

    if os.path.exists(db_path):
        shutil.rmtree(db_path)
    os.makedirs(db_path, exist_ok=True)

    client = chromadb.PersistentClient(path=db_path)
    t0 = time.time()
    col = client.create_collection(name=collection, metadata={"hnsw:space": "cosine"})

    ids = [str(i) for i in range(len(docs))]
    texts = [d["text"] for d in docs]
    metas = [{"source": d["source"]} for d in docs]

    for i in range(0, len(docs), 100):
        j = min(i + 100, len(docs))
        col.add(ids=ids[i:j], embeddings=embeddings[i:j], documents=texts[i:j], metadatas=metas[i:j])
    return time.time() - t0


def write_to_lancedb(model_key, docs, embeddings):
    table = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "lancedb", f"{model_key}_db")

    if os.path.exists(db_path):
        shutil.rmtree(db_path)
    os.makedirs(db_path, exist_ok=True)

    db = lancedb.connect(db_path)
    t0 = time.time()
    data = [
        {"id": i, "vector": emb, "text": d["text"], "source": d["source"]}
        for i, (d, emb) in enumerate(zip(docs, embeddings))
    ]
    db.create_table(table, data=data)
    return time.time() - t0


def write_to_weaviate(model_key, docs, embeddings):
    client = weaviate.connect_to_local()
    try:
        # Weaviate sinif ismi: harfle baslamali, alt cizgi ve sayi olabilir
        collection_name = f"Docs_{model_key}"
        try:
            client.collections.delete(collection_name)
        except Exception:
            pass

        t0 = time.time()
        col = client.collections.create(
            name=collection_name,
            vectorizer_config=Configure.Vectorizer.none(),
            properties=[
                Property(name="text", data_type=DataType.TEXT),
                Property(name="source", data_type=DataType.TEXT),
            ],
        )
        with col.batch.dynamic() as batch:
            for d, emb in zip(docs, embeddings):
                batch.add_object(
                    properties={"text": d["text"], "source": d["source"]},
                    vector=emb,
                )
        return time.time() - t0
    finally:
        client.close()


# ---------------------------------------------------------------
# ARAMA BENCHMARK — her DB kendi algoritmalariyla test edilir
# ---------------------------------------------------------------
def search_milvus(model_key, query_vectors):
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "milvus", f"{model_key}_db.db")
    client = MilvusClient(db_path)
    results = {}

    # 1) Default HNSW
    def default_search():
        return [client.search(collection_name=collection, data=[q], limit=10, output_fields=["text"])
                for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r:
        results["HNSW_default"] = r

    # 2) Farkli limit degerleri
    for limit in [5, 20, 50]:
        def fn(lim=limit):
            return [client.search(collection_name=collection, data=[q], limit=lim, output_fields=["text"])
                    for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r:
            results[f"HNSW_limit{limit}"] = r

    # 3) Batch arama — tum sorgular tek seferde
    def batch():
        return client.search(collection_name=collection, data=query_vectors, limit=10, output_fields=["text"])
    r = measure_time(batch)
    if "error" not in r:
        results["HNSW_batch"] = r

    return results


def search_qdrant(model_key, query_vectors):
    client = QdrantClient(host="localhost", port=6333, timeout=60)
    collection = f"docs_{model_key}"
    results = {}

    def default_search():
        return [client.query_points(collection_name=collection, query=q, limit=10) for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r:
        results["HNSW_default"] = r

    # Exact (brute force) — FLAT eşdegeri
    def exact():
        return [client.query_points(collection_name=collection, query=q, limit=10,
                                    search_params=SearchParams(exact=True))
                for q in query_vectors]
    r = measure_time(exact)
    if "error" not in r:
        results["EXACT"] = r

    # HNSW ef degerleri — recall/hiz dengesi
    for ef in [32, 128, 256]:
        def fn(e=ef):
            return [client.query_points(collection_name=collection, query=q, limit=10,
                                        search_params=SearchParams(hnsw_ef=e))
                    for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r:
            results[f"HNSW_ef{ef}"] = r

    return results


def search_chromadb(model_key, query_vectors):
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "chromadb", f"{model_key}_db")
    client = chromadb.PersistentClient(path=db_path)
    col = client.get_collection(collection)
    results = {}

    def default_search():
        return [col.query(query_embeddings=[q], n_results=10) for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r:
        results["HNSW_default"] = r

    # Farkli n_results
    for n in [5, 20, 50]:
        def fn(nn=n):
            return [col.query(query_embeddings=[q], n_results=nn) for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r:
            results[f"HNSW_n{n}"] = r

    # Batch
    def batch():
        return col.query(query_embeddings=query_vectors, n_results=10)
    r = measure_time(batch)
    if "error" not in r:
        results["HNSW_batch"] = r

    return results


def search_lancedb(model_key, query_vectors):
    table_name = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "lancedb", f"{model_key}_db")
    db = lancedb.connect(db_path)
    table = db.open_table(table_name)
    results = {}

    def default_search():
        return [table.search(q).limit(10).to_pandas() for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r:
        results["VECTOR_default"] = r

    for limit in [5, 20, 50]:
        def fn(lim=limit):
            return [table.search(q).limit(lim).to_pandas() for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r:
            results[f"VECTOR_limit{limit}"] = r

    # Metric karsilastirmasi
    for metric in ["cosine", "L2"]:
        def fn(m=metric):
            return [table.search(q).metric(m).limit(10).to_pandas() for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r:
            results[f"VECTOR_{metric}"] = r

    return results


def search_weaviate(model_key, query_vectors):
    client = weaviate.connect_to_local()
    try:
        col = client.collections.get(f"Docs_{model_key}")
        results = {}

        def default_search():
            return [col.query.near_vector(near_vector=q, limit=10) for q in query_vectors]
        r = measure_time(default_search)
        if "error" not in r:
            results["HNSW_default"] = r

        for limit in [5, 20, 50]:
            def fn(lim=limit):
                return [col.query.near_vector(near_vector=q, limit=lim) for q in query_vectors]
            r = measure_time(fn)
            if "error" not in r:
                results[f"HNSW_limit{limit}"] = r

        # BM25 keyword search
        def bm25():
            return [col.query.bm25(query=qt, limit=10) for qt in TEST_QUERIES]
        r = measure_time(bm25)
        if "error" not in r:
            results["BM25"] = r

        # Hybrid (vektor + BM25)
        for alpha in [0.25, 0.5, 0.75]:
            def fn(a=alpha):
                return [col.query.hybrid(query=qt, vector=q, limit=10, alpha=a)
                        for qt, q in zip(TEST_QUERIES, query_vectors)]
            r = measure_time(fn)
            if "error" not in r:
                results[f"HYBRID_alpha{alpha}"] = r

        return results
    finally:
        client.close()


# ---------------------------------------------------------------
# Sonuclari Excel'e yaz
# ---------------------------------------------------------------
def save_results(results):
    # Once JSON (Excel coksede sonuc kurtarilsin)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n  JSON kaydedildi: {OUTPUT_JSON}")

    try:
        build_excel(results)
        print(f"  Excel kaydedildi: {OUTPUT_XLSX}")
    except Exception as e:
        print(f"  Excel hatasi: {e}")


def build_excel(results):
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 1) OZET
    ws = wb.active
    ws.title = "Ozet"
    ws["A1"] = "BENCHMARK OZETI"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Tarih:"
    ws["B3"] = results["metadata"]["date"]
    ws["A4"] = "Mod:"
    ws["B4"] = results["metadata"]["mode"]
    ws["A5"] = "Belge sayisi:"
    ws["B5"] = results["metadata"]["doc_count"]
    ws["A6"] = "Sorgu sayisi:"
    ws["B6"] = len(TEST_QUERIES)

    ws["A8"] = "MODELLER:"
    ws["A8"].font = Font(bold=True)
    row = 9
    for key, cfg in MODELS.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = cfg["path"]
        ws[f"C{row}"] = f"dim={cfg['dim']}"
        row += 1
    for c in ["A", "B", "C"]:
        ws.column_dimensions[c].width = 40

    # 2) YAZMA SUREIERI
    ws = wb.create_sheet("Yazma Sureleri")
    ws["A1"] = "MODEL x VERITABANI YAZMA SURELERI (saniye)"
    ws["A1"].font = Font(bold=True, size=14)

    dbs = ["milvus", "qdrant", "chromadb", "lancedb", "weaviate"]
    headers = ["Model"] + dbs
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    row = 4
    for model_key in MODELS:
        ws.cell(row=row, column=1, value=model_key).border = border
        for j, db in enumerate(dbs, 2):
            write_data = results["write"].get(db, {}).get(model_key, {})
            if write_data.get("status") == "ok":
                ws.cell(row=row, column=j, value=round(write_data["time"], 3)).border = border
            else:
                ws.cell(row=row, column=j, value="HATA").border = border
            ws.cell(row=row, column=j).alignment = center
        row += 1

    for c in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 16

    # 3) TUM ARAMA SONUCLARI (tek tablo)
    ws = wb.create_sheet("Arama Sonuclari")
    ws["A1"] = "TUM ARAMA TESTLERI (ms)"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Sira", "Veritabani", "Model", "Algoritma", "Ort", "Min", "Max", "Std", "P50", "P95"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    all_rows = []
    for db, db_data in results["search"].items():
        for model_key, algos in db_data.items():
            if not isinstance(algos, dict):
                continue
            for algo, perf in algos.items():
                if "avg_time" not in perf:
                    continue
                all_rows.append({
                    "db": db, "model": model_key, "algo": algo,
                    "avg": perf["avg_time"] * 1000,
                    "min": perf["min_time"] * 1000,
                    "max": perf["max_time"] * 1000,
                    "std": perf["std_time"] * 1000,
                    "p50": perf["p50_time"] * 1000,
                    "p95": perf["p95_time"] * 1000,
                })
    all_rows.sort(key=lambda x: x["avg"])

    row = 4
    for i, r in enumerate(all_rows, 1):
        vals = [i, r["db"], r["model"], r["algo"],
                round(r["avg"], 3), round(r["min"], 3), round(r["max"], 3),
                round(r["std"], 3), round(r["p50"], 3), round(r["p95"], 3)]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.alignment = center
            cell.border = border
            if i <= 3:
                cell.fill = gold
        row += 1

    ws.column_dimensions["A"].width = 6
    for c in ["B", "C", "D"]:
        ws.column_dimensions[c].width = 18
    for c in ["E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[c].width = 10

    # 4) Her DB icin ayri sayfa
    for db in dbs:
        ws = wb.create_sheet(f"{db.upper()}")
        ws["A1"] = f"{db.upper()} — Model x Algoritma (ms ortalama)"
        ws["A1"].font = Font(bold=True, size=14)

        # Hangi algoritmalar var?
        algo_set = set()
        for model_data in results["search"].get(db, {}).values():
            if isinstance(model_data, dict):
                algo_set.update(model_data.keys())
        algos = sorted(algo_set)

        headers = ["Model"] + algos
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center

        row = 4
        for model_key in MODELS:
            ws.cell(row=row, column=1, value=model_key)
            m = results["search"].get(db, {}).get(model_key, {})
            for j, algo in enumerate(algos, 2):
                perf = m.get(algo) if isinstance(m, dict) else None
                if isinstance(perf, dict) and "avg_time" in perf:
                    ws.cell(row=row, column=j, value=round(perf["avg_time"] * 1000, 3))
                else:
                    ws.cell(row=row, column=j, value="-")
                ws.cell(row=row, column=j).alignment = center
            row += 1

        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(algos) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    wb.save(OUTPUT_XLSX)


def print_summary(results):
    print_header("OZET")
    print(f"  Mod: {results['metadata']['mode']}")
    print(f"  Belge: {results['metadata']['doc_count']}")
    print(f"  Sorgu: {len(TEST_QUERIES)}")

    # En hizli 10 arama
    rows = []
    for db, db_data in results["search"].items():
        for model_key, algos in db_data.items():
            if not isinstance(algos, dict):
                continue
            for algo, perf in algos.items():
                if "avg_time" in perf:
                    rows.append((db, model_key, algo, perf["avg_time"] * 1000))
    rows.sort(key=lambda x: x[3])

    print("\n  EN HIZLI 10 ARAMA:")
    print("  " + "-" * 70)
    for i, (db, model, algo, ms) in enumerate(rows[:10], 1):
        print(f"  {i:2}. {db:<10} | {model:<14} | {algo:<22} | {ms:8.3f} ms")


# ---------------------------------------------------------------
# Ana akis
# ---------------------------------------------------------------
def run_benchmark(mode):
    doc_limit = 20 if mode == "test" else None
    docs = load_documents(limit=doc_limit)
    texts = [d["text"] for d in docs]

    # GPU bilgisi
    print_header("SISTEM BILGISI")
    if torch.cuda.is_available():
        gpu = torch.cuda.get_device_name(0)
        vram = torch.cuda.get_device_properties(0).total_memory / 1e9
        print(f"  GPU: {gpu}  ({vram:.2f} GB VRAM)")
    else:
        print("  GPU yok — CPU kullanilacak")

    results = {
        "metadata": {
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "mode": mode,
            "doc_count": len(docs),
            "query_count": len(TEST_QUERIES),
            "models": {k: v["path"] for k, v in MODELS.items()},
        },
        "write": {db: {} for db in ["milvus", "qdrant", "chromadb", "lancedb", "weaviate"]},
        "search": {db: {} for db in ["milvus", "qdrant", "chromadb", "lancedb", "weaviate"]},
    }

    # Sorgu vektorlerini model_key -> vektor seklinde sakla
    # (Arama benchmark asamasinda gerekli olacak)
    query_vectors_by_model = {}

    # ---- ASAMA 1: Her model icin encode + tum DB'lere yaz ----
    for model_key, cfg in MODELS.items():
        print_header(f"MODEL: {model_key}  (dim={cfg['dim']})")

        try:
            doc_embs, query_embs = encode_model(cfg["path"], texts, TEST_QUERIES)
        except Exception as e:
            print(f"  HATA: {model_key} yuklenemedi: {e}")
            for db in results["write"]:
                results["write"][db][model_key] = {"status": "error", "error": str(e)}
            continue

        query_vectors_by_model[model_key] = query_embs

        # Her veritabanina sirayla yaz
        write_funcs = {
            "milvus":   write_to_milvus,
            "qdrant":   write_to_qdrant,
            "chromadb": write_to_chromadb,
            "lancedb":  write_to_lancedb,
            "weaviate": write_to_weaviate,
        }

        for db_name, fn in write_funcs.items():
            print(f"  -> {db_name} yaziliyor...")
            try:
                t = fn(model_key, docs, doc_embs)
                results["write"][db_name][model_key] = {
                    "status": "ok", "time": t, "record_count": len(docs), "dim": cfg["dim"],
                }
                print(f"     OK: {t:.2f}s")
            except Exception as e:
                results["write"][db_name][model_key] = {"status": "error", "error": str(e)}
                print(f"     HATA: {e}")

        # Belge embeddinglerini bellekten at — siradaki modelin yeri olsun
        del doc_embs
        free_memory()

        # Her modelden sonra ara checkpoint
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2, ensure_ascii=False, default=str)

    # ---- ASAMA 2: Her DB'de her model icin arama benchmark ----
    print_header("ARAMA BENCHMARK BASLIYOR")

    search_funcs = {
        "milvus":   search_milvus,
        "qdrant":   search_qdrant,
        "chromadb": search_chromadb,
        "lancedb":  search_lancedb,
        "weaviate": search_weaviate,
    }

    for db_name, fn in search_funcs.items():
        print_header(f"ARAMA: {db_name.upper()}")
        for model_key in MODELS:
            qvs = query_vectors_by_model.get(model_key)
            if qvs is None:
                continue
            print(f"  {model_key}...")
            try:
                algo_results = fn(model_key, qvs)
                results["search"][db_name][model_key] = algo_results
                for algo, perf in algo_results.items():
                    print(f"    {algo:<22} {perf['avg_time']*1000:8.3f} ms")
            except Exception as e:
                results["search"][db_name][model_key] = {"error": str(e)}
                print(f"    HATA: {e}")

        # Her DB sonrasi checkpoint
        with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
            json.dump(results, f, indent=2, ensure_ascii=False, default=str)

    # ---- SONUC ----
    save_results(results)
    print_summary(results)


def main():
    parser = argparse.ArgumentParser(description="Vektor DB Multi-Model Benchmark")
    parser.add_argument(
        "--mode",
        choices=["test", "full"],
        default="test",
        help="test = 20 belge ile hizli test, full = tum belgelerle tam benchmark",
    )
    args = parser.parse_args()

    print("=" * 70)
    print(f"  VEKTOR DATABASE BENCHMARK — mod: {args.mode.upper()}")
    print("=" * 70)

    try:
        run_benchmark(args.mode)
    except KeyboardInterrupt:
        print("\n  Kullanici iptali — cikiliyor")


if __name__ == "__main__":
    main()
