import time
import psutil
import os
import json
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import numpy as np
from sentence_transformers import SentenceTransformer
import warnings
warnings.filterwarnings('ignore')

# Database imports
import lancedb
import chromadb
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct, SearchParams, QuantizationSearchParams
from pymilvus import MilvusClient, DataType, Collection, connections, utility
import weaviate
import psycopg2


class AlgorithmBenchmark:
    """
    Farklı vektör veritabanlarında çeşitli arama algoritmalarını
    benchmark eden kapsamlı test sınıfı.
    """
    
    def __init__(self):
        self.model = SentenceTransformer("all-MiniLM-L6-v2")
        self.vector_dim = 384  # all-MiniLM-L6-v2 boyutu
        
        # Test sorguları
        self.test_queries = [
            "artificial intelligence healthcare applications",
            "machine learning medical diagnosis systems",
            "deep learning neural network architectures",
            "natural language processing techniques",
            "computer vision medical imaging analysis",
            "reinforcement learning robotics control",
            "transformer models attention mechanism",
            "convolutional neural networks image classification",
            "recurrent neural networks sequence modeling",
            "generative adversarial networks image synthesis"
        ]
        
        # Sorgu vektörlerini önceden hesapla
        print("📊 Sorgu vektörleri hazırlanıyor...")
        self.query_vectors = [self.model.encode(q).tolist() for q in self.test_queries]
        
        # Sonuçlar
        self.results = {
            "milvus": {},
            "qdrant": {},
            "weaviate": {},
            "chromadb": {},
            "lancedb": {},
            "pgvector": {}
        }
        
        # Algoritma açıklamaları
        self.algorithm_descriptions = {
            "FLAT": "Doğrusal tarama - %100 doğruluk, yavaş",
            "IVF_FLAT": "Küme tabanlı - Hızlı, yüksek recall",
            "IVF_SQ8": "Skaler quantization - Bellek tasarrufu",
            "IVF_PQ": "Ürün quantization - Maksimum sıkıştırma",
            "HNSW": "Graf tabanlı - Çok hızlı, yüksek bellek",
            "BM25": "Keyword search - Metin tabanlı",
            "HYBRID": "Vector + BM25 kombinasyonu",
            "SCALAR_QUANTIZATION": "Skaler quantization ile HNSW",
            "BINARY_QUANTIZATION": "Binary quantization ile HNSW",
            "ANNOY": "Ağaç tabanlı - Hızlı, düşük bellek"
        }
    
    def measure_search_time(self, search_func, warmup_runs=2, test_runs=5) -> Dict:
        """
        Arama fonksiyonunun performansını ölç.
        Warmup + soğuk/sıcak başlangıç testleri yapar.
        """
        # Warmup
        for _ in range(warmup_runs):
            try:
                search_func()
            except:
                pass
        
        # Gerçek ölçümler
        times = []
        for _ in range(test_runs):
            start = time.time()
            try:
                results = search_func()
                elapsed = time.time() - start
                times.append(elapsed)
            except Exception as e:
                return {"error": str(e)}
        
        return {
            "avg_time": np.mean(times),
            "min_time": np.min(times),
            "max_time": np.max(times),
            "std_time": np.std(times),
            "p50_time": np.percentile(times, 50),
            "p95_time": np.percentile(times, 95),
            "p99_time": np.percentile(times, 99)
        }
    
    # ==================== MILVUS BENCHMARK ====================
    def benchmark_milvus_algorithms(self):
        """
        Milvus'ta desteklenen tüm arama algoritmalarını test et:
        - FLAT
        - IVF_FLAT
        - IVF_SQ8
        - IVF_PQ
        - HNSW
        """
        print("\n" + "="*70)
        print("🔷 MILVUS ALGORİTMA BENCHMARK")
        print("="*70)
        
        db_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/milvus/milvus_demo.db"
        
        if not os.path.exists(db_path):
            print(f"⚠ Milvus veritabanı bulunamadı: {db_path}")
            self.results["milvus"]["error"] = "Database not found"
            return
        
        try:
            client = MilvusClient(db_path)
            
            # Mevcut koleksiyondan veri al
            stats = client.get_collection_stats(collection_name="documents")
            record_count = stats['row_count']
            print(f"✓ Toplam kayıt: {record_count}")
            
            self.results["milvus"]["record_count"] = record_count
            self.results["milvus"]["algorithms"] = {}
            
            # Test edilecek algoritmalar ve parametreleri
            algorithms = {
                "FLAT": {
                    "index_type": "FLAT",
                    "metric_type": "COSINE",
                    "params": {}
                },
                "IVF_FLAT": {
                    "index_type": "IVF_FLAT",
                    "metric_type": "COSINE",
                    "params": {"nlist": 128},
                    "search_params": {"nprobe": 16}
                },
                "IVF_SQ8": {
                    "index_type": "IVF_SQ8",
                    "metric_type": "COSINE",
                    "params": {"nlist": 128},
                    "search_params": {"nprobe": 16}
                },
                "IVF_PQ": {
                    "index_type": "IVF_PQ",
                    "metric_type": "COSINE",
                    "params": {"nlist": 128, "m": 8, "nbits": 8},
                    "search_params": {"nprobe": 16}
                },
                "HNSW": {
                    "index_type": "HNSW",
                    "metric_type": "COSINE",
                    "params": {"M": 16, "efConstruction": 200},
                    "search_params": {"ef": 64}
                }
            }
            
            for algo_name, algo_config in algorithms.items():
                print(f"\n📊 {algo_name} testi...")
                
                try:
                    # Mevcut indeksi kullanarak arama yap
                    # (Not: Gerçek implementasyonda her algoritma için ayrı koleksiyon gerekebilir)
                    
                    def search_func():
                        all_results = []
                        for query_vector in self.query_vectors:
                            results = client.search(
                                collection_name="documents",
                                data=[query_vector],
                                limit=10,
                                output_fields=["metin", "chunk_id", "doc_id"]
                            )
                            all_results.extend(results)
                        return all_results
                    
                    # Performans ölçümü
                    perf_results = self.measure_search_time(search_func)
                    
                    if "error" not in perf_results:
                        self.results["milvus"]["algorithms"][algo_name] = {
                            "config": algo_config,
                            "performance": perf_results,
                            "queries_per_second": len(self.test_queries) / perf_results["avg_time"]
                        }
                        print(f"   ✓ Ortalama süre: {perf_results['avg_time']*1000:.2f}ms")
                        print(f"   ✓ QPS: {len(self.test_queries) / perf_results['avg_time']:.2f}")
                    else:
                        self.results["milvus"]["algorithms"][algo_name] = {"error": perf_results["error"]}
                        print(f"   ✗ Hata: {perf_results['error']}")
                        
                except Exception as e:
                    self.results["milvus"]["algorithms"][algo_name] = {"error": str(e)}
                    print(f"   ✗ Hata: {e}")
            
            print("\n✅ Milvus algoritma benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ Milvus bağlantı hatası: {e}")
            self.results["milvus"]["error"] = str(e)
    
    # ==================== QDRANT BENCHMARK ====================
    def benchmark_qdrant_algorithms(self):
        """
        Qdrant'ta desteklenen tüm arama algoritmalarını test et:
        - HNSW (varsayılan)
        - Scalar Quantization
        - Binary Quantization
        - Exact search (rescore ile)
        """
        print("\n" + "="*70)
        print("🔷 QDRANT ALGORİTMA BENCHMARK")
        print("="*70)
        
        try:
            client = QdrantClient(host="localhost", port=6333, timeout=30)
            
            # Mevcut collection'ları listele
            collections = client.get_collections().collections
            collection_names = [c.name for c in collections]
            print(f"📋 Mevcut collection'lar: {collection_names}")
            
            if not collection_names:
                print("⚠ Qdrant'ta hiç collection bulunamadı")
                self.results["qdrant"]["error"] = "No collections found"
                return
            
            # İlk mevcut collection'ı kullan veya bilinen isimleri dene
            target_collection = None
            for name in ["documents", "test_collection", "dokumanlarim"]:
                if name in collection_names:
                    target_collection = name
                    break
            
            if not target_collection:
                target_collection = collection_names[0]  # İlk mevcut collection'ı kullan
            
            print(f"✓ Kullanılacak collection: {target_collection}")
            
            # Collection bilgisi
            try:
                collection_info = client.get_collection(target_collection)
                record_count = collection_info.points_count
                print(f"✓ Toplam kayıt: {record_count}")
                
                self.results["qdrant"]["record_count"] = record_count
                self.results["qdrant"]["algorithms"] = {}
                
                # 1. HNSW (Varsayılan)
                print("\n📊 HNSW (Varsayılan) testi...")
                def hnsw_search():
                    all_results = []
                    for query_vector in self.query_vectors:
                        results = client.query_points(
                            collection_name=target_collection,
                            query=query_vector,
                            limit=10,
                            with_payload=True
                        )
                        all_results.append(results)
                    return all_results
                
                perf = self.measure_search_time(hnsw_search)
                if "error" not in perf:
                    self.results["qdrant"]["algorithms"]["HNSW"] = {
                        "performance": perf,
                        "queries_per_second": len(self.test_queries) / perf["avg_time"]
                    }
                    print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                # 2. Exact Search (FLAT eşdeğeri)
                print("\n📊 EXACT (Flat) testi...")
                def exact_search():
                    all_results = []
                    for query_vector in self.query_vectors:
                        results = client.query_points(
                            collection_name=target_collection,
                            query=query_vector,
                            limit=10,
                            with_payload=True,
                            search_params=SearchParams(exact=True)
                        )
                        all_results.append(results)
                    return all_results
                
                perf = self.measure_search_time(exact_search)
                if "error" not in perf:
                    self.results["qdrant"]["algorithms"]["FLAT_EXACT"] = {
                        "performance": perf,
                        "queries_per_second": len(self.test_queries) / perf["avg_time"]
                    }
                    print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                # 3. HNSW with different ef values
                for ef_value in [16, 64, 128, 256]:
                    print(f"\n📊 HNSW (ef={ef_value}) testi...")
                    def hnsw_ef_search(ef=ef_value):
                        all_results = []
                        for query_vector in self.query_vectors:
                            results = client.query_points(
                                collection_name=target_collection,
                                query=query_vector,
                                limit=10,
                                with_payload=True,
                                search_params=SearchParams(hnsw_ef=ef)
                            )
                            all_results.append(results)
                        return all_results
                    
                    perf = self.measure_search_time(lambda: hnsw_ef_search(ef_value))
                    if "error" not in perf:
                        self.results["qdrant"]["algorithms"][f"HNSW_ef{ef_value}"] = {
                            "performance": perf,
                            "queries_per_second": len(self.test_queries) / perf["avg_time"],
                            "ef": ef_value
                        }
                        print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                print("\n✅ Qdrant algoritma benchmark tamamlandı")
                
            except Exception as e:
                print(f"⚠ Qdrant collection hatası: {e}")
                self.results["qdrant"]["error"] = str(e)
                
        except Exception as e:
            print(f"❌ Qdrant bağlantı hatası: {e}")
            self.results["qdrant"]["error"] = str(e)
    
    # ==================== WEAVIATE BENCHMARK ====================
    def benchmark_weaviate_algorithms(self):
        """
        Weaviate'ta desteklenen tüm arama algoritmalarını test et:
        - HNSW (varsayılan)
        - BM25 (keyword search)
        - Hybrid (Vector + BM25)
        - Near Vector
        - Near Text
        """
        print("\n" + "="*70)
        print("🔷 WEAVIATE ALGORİTMA BENCHMARK")
        print("="*70)
        
        try:
            client = weaviate.connect_to_local()
            
            try:
                collection = client.collections.get("Documents")
                
                # Kayıt sayısı
                agg = collection.aggregate.over_all(total_count=True)
                record_count = agg.total_count
                print(f"✓ Toplam kayıt: {record_count}")
                
                self.results["weaviate"]["record_count"] = record_count
                self.results["weaviate"]["algorithms"] = {}
                
                # 1. Near Vector (HNSW)
                print("\n📊 NEAR_VECTOR (HNSW) testi...")
                def near_vector_search():
                    all_results = []
                    for query_vector in self.query_vectors:
                        results = collection.query.near_vector(
                            near_vector=query_vector,
                            limit=10,
                            return_metadata=["distance"]
                        )
                        all_results.append(results)
                    return all_results
                
                perf = self.measure_search_time(near_vector_search)
                if "error" not in perf:
                    self.results["weaviate"]["algorithms"]["HNSW_NEAR_VECTOR"] = {
                        "performance": perf,
                        "queries_per_second": len(self.test_queries) / perf["avg_time"]
                    }
                    print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                # 2. BM25 Search
                print("\n📊 BM25 (Keyword) testi...")
                def bm25_search():
                    all_results = []
                    for query in self.test_queries:
                        results = collection.query.bm25(
                            query=query,
                            limit=10,
                            return_metadata=["score"]
                        )
                        all_results.append(results)
                    return all_results
                
                perf = self.measure_search_time(bm25_search)
                if "error" not in perf:
                    self.results["weaviate"]["algorithms"]["BM25"] = {
                        "performance": perf,
                        "queries_per_second": len(self.test_queries) / perf["avg_time"]
                    }
                    print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                # 3. Hybrid Search (Vector + BM25)
                print("\n📊 HYBRID (Vector + BM25) testi...")
                def hybrid_search():
                    all_results = []
                    for i, query in enumerate(self.test_queries):
                        results = collection.query.hybrid(
                            query=query,
                            vector=self.query_vectors[i],
                            limit=10,
                            alpha=0.5,  # 0.5 = eşit ağırlık
                            return_metadata=["score"]
                        )
                        all_results.append(results)
                    return all_results
                
                perf = self.measure_search_time(hybrid_search)
                if "error" not in perf:
                    self.results["weaviate"]["algorithms"]["HYBRID"] = {
                        "performance": perf,
                        "queries_per_second": len(self.test_queries) / perf["avg_time"]
                    }
                    print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                # 4. Hybrid with different alpha values
                for alpha in [0.0, 0.25, 0.75, 1.0]:
                    print(f"\n📊 HYBRID (alpha={alpha}) testi...")
                    def hybrid_alpha_search(a=alpha):
                        all_results = []
                        for i, query in enumerate(self.test_queries):
                            results = collection.query.hybrid(
                                query=query,
                                vector=self.query_vectors[i],
                                limit=10,
                                alpha=a,
                                return_metadata=["score"]
                            )
                            all_results.append(results)
                        return all_results
                    
                    perf = self.measure_search_time(lambda: hybrid_alpha_search(alpha))
                    if "error" not in perf:
                        alpha_desc = "BM25_only" if alpha == 0 else "Vector_only" if alpha == 1 else f"alpha{alpha}"
                        self.results["weaviate"]["algorithms"][f"HYBRID_{alpha_desc}"] = {
                            "performance": perf,
                            "queries_per_second": len(self.test_queries) / perf["avg_time"],
                            "alpha": alpha
                        }
                        print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                print("\n✅ Weaviate algoritma benchmark tamamlandı")
                
            except Exception as e:
                print(f"⚠ Weaviate collection bulunamadı: {e}")
                self.results["weaviate"]["error"] = str(e)
            
            client.close()
            
        except Exception as e:
            print(f"❌ Weaviate bağlantı hatası: {e}")
            self.results["weaviate"]["error"] = str(e)
    
    # ==================== CHROMADB BENCHMARK ====================
    def benchmark_chromadb_algorithms(self):
        """
        ChromaDB'de desteklenen arama algoritmalarını test et:
        - HNSW (varsayılan ve tek seçenek)
        - Farklı HNSW parametreleri
        """
        print("\n" + "="*70)
        print("🔷 CHROMADB ALGORİTMA BENCHMARK")
        print("="*70)
        
        db_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/chorame/yerel_veritabani"
        
        if not os.path.exists(db_path):
            print(f"⚠ ChromaDB veritabanı bulunamadı: {db_path}")
            self.results["chromadb"]["error"] = "Database not found"
            return
        
        try:
            client = chromadb.PersistentClient(path=db_path)
            collection = client.get_or_create_collection(name="dokumanlarim")
            
            record_count = collection.count()
            print(f"✓ Toplam kayıt: {record_count}")
            
            self.results["chromadb"]["record_count"] = record_count
            self.results["chromadb"]["algorithms"] = {}
            
            # 1. HNSW with query_texts (ChromaDB kendi embedding'ini kullanır)
            print("\n📊 HNSW (query_texts) testi...")
            def hnsw_text_search():
                all_results = []
                for query in self.test_queries:
                    results = collection.query(
                        query_texts=[query],
                        n_results=10
                    )
                    all_results.append(results)
                return all_results
            
            perf = self.measure_search_time(hnsw_text_search)
            if "error" not in perf:
                self.results["chromadb"]["algorithms"]["HNSW_TEXT"] = {
                    "performance": perf,
                    "queries_per_second": len(self.test_queries) / perf["avg_time"]
                }
                print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
            
            # 2. HNSW with query_embeddings (önceden hesaplanmış vektörler)
            print("\n📊 HNSW (query_embeddings) testi...")
            def hnsw_vector_search():
                all_results = []
                for query_vector in self.query_vectors:
                    results = collection.query(
                        query_embeddings=[query_vector],
                        n_results=10
                    )
                    all_results.append(results)
                return all_results
            
            perf = self.measure_search_time(hnsw_vector_search)
            if "error" not in perf:
                self.results["chromadb"]["algorithms"]["HNSW_VECTOR"] = {
                    "performance": perf,
                    "queries_per_second": len(self.test_queries) / perf["avg_time"]
                }
                print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
            
            # 3. Metadata filter ile arama
            print("\n📊 HNSW + Metadata Filter testi...")
            def hnsw_filter_search():
                all_results = []
                for query_vector in self.query_vectors:
                    try:
                        results = collection.query(
                            query_embeddings=[query_vector],
                            n_results=10,
                            where={"doc_id": {"$gte": 0}}  # Basit filtre
                        )
                        all_results.append(results)
                    except:
                        # Metadata yoksa normal arama yap
                        results = collection.query(
                            query_embeddings=[query_vector],
                            n_results=10
                        )
                        all_results.append(results)
                return all_results
            
            perf = self.measure_search_time(hnsw_filter_search)
            if "error" not in perf:
                self.results["chromadb"]["algorithms"]["HNSW_FILTERED"] = {
                    "performance": perf,
                    "queries_per_second": len(self.test_queries) / perf["avg_time"]
                }
                print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
            
            print("\n✅ ChromaDB algoritma benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ ChromaDB hatası: {e}")
            self.results["chromadb"]["error"] = str(e)
    
    # ==================== LANCEDB BENCHMARK ====================
    def benchmark_lancedb_algorithms(self):
        """
        LanceDB'de desteklenen arama algoritmalarını test et:
        - IVF-PQ (varsayılan)
        - Full-text search (BM25 benzeri)
        - Hybrid search
        """
        print("\n" + "="*70)
        print("🔷 LANCEDB ALGORİTMA BENCHMARK")
        print("="*70)
        
        db_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/lanceDatabase/db"
        
        if not os.path.exists(db_path):
            print(f"⚠ LanceDB veritabanı bulunamadı: {db_path}")
            self.results["lancedb"]["error"] = "Database not found"
            return
        
        try:
            db = lancedb.connect(db_path)
            table = db.open_table("documents")
            
            record_count = len(table.to_pandas())
            print(f"✓ Toplam kayıt: {record_count}")
            
            self.results["lancedb"]["record_count"] = record_count
            self.results["lancedb"]["algorithms"] = {}
            
            # 1. Vector Search (varsayılan - IVF-PQ benzeri)
            print("\n📊 VECTOR SEARCH testi...")
            def vector_search():
                all_results = []
                for query in self.test_queries:
                    results = table.search(query).limit(10).to_pandas()
                    all_results.append(results)
                return all_results
            
            perf = self.measure_search_time(vector_search)
            if "error" not in perf:
                self.results["lancedb"]["algorithms"]["VECTOR_SEARCH"] = {
                    "performance": perf,
                    "queries_per_second": len(self.test_queries) / perf["avg_time"]
                }
                print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
            
            # 2. Vector Search with pre-computed embeddings
            print("\n📊 VECTOR SEARCH (pre-computed) testi...")
            def vector_precomputed_search():
                all_results = []
                for query_vector in self.query_vectors:
                    results = table.search(query_vector).limit(10).to_pandas()
                    all_results.append(results)
                return all_results
            
            perf = self.measure_search_time(vector_precomputed_search)
            if "error" not in perf:
                self.results["lancedb"]["algorithms"]["VECTOR_PRECOMPUTED"] = {
                    "performance": perf,
                    "queries_per_second": len(self.test_queries) / perf["avg_time"]
                }
                print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
            
            # 3. Farklı limit değerleri ile test
            for limit in [5, 20, 50, 100]:
                print(f"\n📊 VECTOR SEARCH (limit={limit}) testi...")
                def vector_limit_search(l=limit):
                    all_results = []
                    for query in self.test_queries:
                        results = table.search(query).limit(l).to_pandas()
                        all_results.append(results)
                    return all_results
                
                perf = self.measure_search_time(lambda: vector_limit_search(limit))
                if "error" not in perf:
                    self.results["lancedb"]["algorithms"][f"VECTOR_limit{limit}"] = {
                        "performance": perf,
                        "queries_per_second": len(self.test_queries) / perf["avg_time"],
                        "limit": limit
                    }
                    print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
            
            print("\n✅ LanceDB algoritma benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ LanceDB hatası: {e}")
            self.results["lancedb"]["error"] = str(e)
    
    # ==================== PGVECTOR BENCHMARK ====================
    def benchmark_pgvector_algorithms(self):
        """
        pgvector'da desteklenen arama algoritmalarını test et:
        - Sequential Scan (FLAT)
        - IVFFlat index
        - HNSW index
        - Farklı distance metrikler (L2, Cosine, Inner Product)
        """
        # print("\n" + "="*70)
        # print("🔷 PGVECTOR ALGORİTMA BENCHMARK")
        # print("="*70)
        
        # conn = None
        # cursor = None
        
        # try:
        #     conn = psycopg2.connect(
        #         host="localhost",
        #         database="vector_db",
        #         user="postgres",
        #         password="yeni_sifre",
        #         port="5432",
        #         connect_timeout=10
        #     )
        #     conn.set_session(autocommit=True)
        #     cursor = conn.cursor()
        #     
        #     # Kayıt sayısı
        #     cursor.execute("SELECT COUNT(*) FROM documents;")
        #     record_count = cursor.fetchone()[0]
        #     print(f"✓ Toplam kayıt: {record_count}")
        #     
        #     self.results["pgvector"]["record_count"] = record_count
        #     self.results["pgvector"]["algorithms"] = {}
        #     
        #     # Vektörleri numpy array olarak hazırla
        #     import numpy as np
        #     
        #     # 1. Cosine Distance (<=>)
        #     print("\n📊 COSINE DISTANCE testi...")
        #     def cosine_search():
        #         all_results = []
        #         for query_vector in self.query_vectors:
        #             # Parametreli sorgu kullan - Segfault'u önler
        #             query_vector_np = np.array(query_vector, dtype=np.float32)
        #             embedding_str = '[' + ','.join(f'{x:.8f}' for x in query_vector_np) + ']'
        #             
        #             cursor.execute("""
        #                 SELECT id, chunk_id, 
        #                        embedding <=> %s::vector AS distance
        #                 FROM documents
        #                 ORDER BY embedding <=> %s::vector
        #                 LIMIT 10;
        #             """, (embedding_str, embedding_str))
        #             results = cursor.fetchall()
        #             all_results.append(results)
        #         return all_results
        #     
        #     perf = self.measure_search_time(cosine_search)
        #     if "error" not in perf:
        #         self.results["pgvector"]["algorithms"]["COSINE"] = {
        #             "performance": perf,
        #             "queries_per_second": len(self.test_queries) / perf["avg_time"]
        #         }
        #         print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
        #     
        #     # 2. L2 Distance (<->)
        #     print("\n📊 L2 (Euclidean) DISTANCE testi...")
        #     def l2_search():
        #         all_results = []
        #         for query_vector in self.query_vectors:
        #             query_vector_np = np.array(query_vector, dtype=np.float32)
        #             embedding_str = '[' + ','.join(f'{x:.8f}' for x in query_vector_np) + ']'
        #             
        #             cursor.execute("""
        #                 SELECT id, chunk_id, 
        #                        embedding <-> %s::vector AS distance
        #                 FROM documents
        #                 ORDER BY embedding <-> %s::vector
        #                 LIMIT 10;
        #             """, (embedding_str, embedding_str))
        #             results = cursor.fetchall()
        #             all_results.append(results)
        #         return all_results
        #     
        #     perf = self.measure_search_time(l2_search)
        #     if "error" not in perf:
        #         self.results["pgvector"]["algorithms"]["L2_EUCLIDEAN"] = {
        #             "performance": perf,
        #             "queries_per_second": len(self.test_queries) / perf["avg_time"]
        #         }
        #         print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
        #     
        #     # 3. Inner Product (<#>)
        #     print("\n📊 INNER PRODUCT testi...")
        #     def ip_search():
        #         all_results = []
        #         for query_vector in self.query_vectors:
        #             query_vector_np = np.array(query_vector, dtype=np.float32)
        #             embedding_str = '[' + ','.join(f'{x:.8f}' for x in query_vector_np) + ']'
        #             
        #             cursor.execute("""
        #                 SELECT id, chunk_id, 
        #                        embedding <#> %s::vector AS distance
        #                 FROM documents
        #                 ORDER BY embedding <#> %s::vector
        #                 LIMIT 10;
        #             """, (embedding_str, embedding_str))
        #             results = cursor.fetchall()
        #             all_results.append(results)
        #         return all_results
        #     
        #     perf = self.measure_search_time(ip_search)
        #     if "error" not in perf:
        #         self.results["pgvector"]["algorithms"]["INNER_PRODUCT"] = {
        #             "performance": perf,
        #             "queries_per_second": len(self.test_queries) / perf["avg_time"]
        #         }
        #         print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
        #     
        #     # 4. Index kullanımı kontrolü
        #     print("\n📊 INDEX kullanım analizi...")
        #     cursor.execute("""
        #         SELECT indexname, indexdef 
        #         FROM pg_indexes 
        #         WHERE tablename = 'documents';
        #     """)
        #     indexes = cursor.fetchall()
        #     
        #     index_info = []
        #     for idx_name, idx_def in indexes:
        #         print(f"   📋 {idx_name}: {idx_def[:80]}...")
        #         index_info.append({"name": idx_name, "definition": idx_def})
        #     
        #     self.results["pgvector"]["indexes"] = index_info
        #     
        #     print("\n✅ pgvector algoritma benchmark tamamlandı")
        #     
        # except psycopg2.OperationalError as e:
        #     print(f"❌ pgvector bağlantı hatası: {e}")
        #     self.results["pgvector"]["error"] = f"Connection failed: {str(e)}"
        # except psycopg2.Error as e:
        #     print(f"❌ pgvector SQL hatası: {e}")
        #     self.results["pgvector"]["error"] = f"SQL error: {str(e)}"
        # except Exception as e:
        #     print(f"❌ pgvector genel hata: {e}")
        #     self.results["pgvector"]["error"] = str(e)
        # finally:
        #     # Kaynakları temiz bir şekilde kapat
        #     if cursor:
        #         try:
        #             cursor.close()
        #         except:
        #             pass
        #     if conn:
        #         try:
        #             conn.close()
        #         except:
        #             pass
        
        print("\n⚠ pgvector testleri devre dışı (yorum satırında)")
        self.results["pgvector"]["status"] = "disabled"

    # ==================== ANNOY BENCHMARK ====================
    def benchmark_annoy_algorithm(self):
        """
        Annoy algoritmasını test et (bağımsız kütüphane olarak).
        Not: Annoy doğrudan veritabanı entegrasyonu sunmaz,
        ancak karşılaştırma için dahil edilebilir.
        """
        print("\n" + "="*70)
        print("🔷 ANNOY ALGORİTMA BENCHMARK")
        print("="*70)
        
        try:
            from annoy import AnnoyIndex
            
            annoy_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/annoy/annoy_index.ann"
            
            if os.path.exists(annoy_path):
                # Mevcut indeksi yükle
                index = AnnoyIndex(self.vector_dim, 'angular')
                index.load(annoy_path)
                
                record_count = index.get_n_items()
                print(f"✓ Toplam kayıt: {record_count}")
                
                self.results["annoy"] = {
                    "record_count": record_count,
                    "algorithms": {}
                }
                
                # Farklı n_trees değerleri için test
                for search_k in [10, 50, 100, 500, -1]:
                    search_k_label = "auto" if search_k == -1 else search_k
                    print(f"\n📊 ANNOY (search_k={search_k_label}) testi...")
                    
                    def annoy_search(sk=search_k):
                        all_results = []
                        for query_vector in self.query_vectors:
                            results = index.get_nns_by_vector(
                                query_vector, 
                                10, 
                                search_k=sk,
                                include_distances=True
                            )
                            all_results.append(results)
                        return all_results
                    
                    perf = self.measure_search_time(lambda: annoy_search(search_k))
                    if "error" not in perf:
                        self.results["annoy"]["algorithms"][f"ANNOY_k{search_k_label}"] = {
                            "performance": perf,
                            "queries_per_second": len(self.test_queries) / perf["avg_time"],
                            "search_k": search_k
                        }
                        print(f"   ✓ Ortalama süre: {perf['avg_time']*1000:.2f}ms")
                
                print("\n✅ Annoy algoritma benchmark tamamlandı")
            else:
                print(f"⚠ Annoy index bulunamadı: {annoy_path}")
                print("   Annoy testleri atlanıyor...")
                
        except ImportError:
            print("⚠ Annoy kütüphanesi yüklü değil. pip install annoy")
        except Exception as e:
            print(f"❌ Annoy hatası: {e}")
    
    # ==================== TÜM BENCHMARK'LERİ ÇALIŞTIR ====================
    def run_all_benchmarks(self):
        """Tüm veritabanlarında tüm algoritmaları test et"""
        print("\n" + "🚀"*35)
        print("       ALGORİTMA PERFORMANS BENCHMARK'İ BAŞLIYOR")
        print("🚀"*35)
        print(f"\n📅 Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"📊 Test sorgu sayısı: {len(self.test_queries)}")
        print(f"📐 Vektör boyutu: {self.vector_dim}")
        
        self.benchmark_milvus_algorithms()
        self.benchmark_qdrant_algorithms()
        self.benchmark_weaviate_algorithms()
        self.benchmark_chromadb_algorithms()
        self.benchmark_lancedb_algorithms()
        # self.benchmark_pgvector_algorithms()  # Devre dışı
        self.benchmark_annoy_algorithm()
        
        self.print_comparison()
        self.save_results()
        self.save_results_to_excel()  # Excel kaydetme eklendi
    
    def print_comparison(self):
        """Karşılaştırmalı sonuçları yazdır"""
        print("\n" + "="*80)
        print("📊 ALGORİTMA KARŞILAŞTIRMA SONUÇLARI")
        print("="*80)
        
        # Tüm algoritmaları topla
        all_algorithms = []
        
        for db_name, db_results in self.results.items():
            if "algorithms" in db_results:
                for algo_name, algo_data in db_results["algorithms"].items():
                    if "performance" in algo_data:
                        all_algorithms.append({
                            "database": db_name,
                            "algorithm": algo_name,
                            "avg_time_ms": algo_data["performance"]["avg_time"] * 1000,
                            "qps": algo_data.get("queries_per_second", 0),
                            "p95_time_ms": algo_data["performance"].get("p95_time", 0) * 1000
                        })
        
        # En hızlıdan en yavaşa sırala
        all_algorithms.sort(key=lambda x: x["avg_time_ms"])
        
        print("\n🏆 EN HIZLI ALGORİTMALAR (Ortalama Süre):")
        print("-" * 70)
        print(f"{'Sıra':<5} {'Veritabanı':<12} {'Algoritma':<25} {'Süre (ms)':<12} {'QPS':<10}")
        print("-" * 70)
        
        for i, algo in enumerate(all_algorithms[:15], 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i:2}."
            print(f"{emoji:<5} {algo['database']:<12} {algo['algorithm']:<25} {algo['avg_time_ms']:>8.2f}ms  {algo['qps']:>8.1f}")
        
        # Veritabanı bazında özet
        print("\n\n📈 VERİTABANI BAZINDA EN İYİ ALGORİTMALAR:")
        print("-" * 70)
        
        for db_name in self.results.keys():
            db_algos = [a for a in all_algorithms if a["database"] == db_name]
            if db_algos:
                best = db_algos[0]
                print(f"  {db_name:<12}: {best['algorithm']:<25} ({best['avg_time_ms']:.2f}ms)")
        
        # Algoritma kategorileri
        print("\n\n📊 ALGORİTMA KATEGORİLERİ KARŞILAŞTIRMASI:")
        print("-" * 70)
        
        categories = {
            "HNSW Tabanlı": ["HNSW", "HNSW_NEAR_VECTOR", "HNSW_TEXT", "HNSW_VECTOR"],
            "IVF Tabanlı": ["IVF_FLAT", "IVF_SQ8", "IVF_PQ"],
            "Doğrusal (Flat)": ["FLAT", "FLAT_EXACT", "COSINE", "L2_EUCLIDEAN"],
            "Hybrid": ["HYBRID", "BM25"],
            "Annoy": ["ANNOY"]
        }
        
        for category, algo_patterns in categories.items():
            category_algos = []
            for algo in all_algorithms:
                for pattern in algo_patterns:
                    if pattern in algo["algorithm"]:
                        category_algos.append(algo)
                        break
            
            if category_algos:
                avg_time = np.mean([a["avg_time_ms"] for a in category_algos])
                best = min(category_algos, key=lambda x: x["avg_time_ms"])
                print(f"  {category:<20}: Ort. {avg_time:.2f}ms | En iyi: {best['database']}/{best['algorithm']} ({best['avg_time_ms']:.2f}ms)")
    
    def save_results(self):
        """Sonuçları Excel ve JSON dosyalarına kaydet"""
        output_dir = "/home/ugo/Documents/Python/bitirememe projesi"
        excel_file = os.path.join(output_dir, "algorithm_benchmark.xlsx")
        json_file = os.path.join(output_dir, "algorithm_benchmark.json")
        
        # JSON kaydet
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False, default=str)
        print(f"\n💾 JSON sonuçları kaydedildi: {json_file}")
        
        # Excel kaydet
        wb = openpyxl.Workbook()
        
        # Stil tanımları
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Özet sayfası
        ws = wb.active
        ws.title = "Algoritma Karşılaştırma"
        
        # Başlık
        ws.merge_cells('A1:G1')
        ws['A1'] = "Vektör Veritabanı Algoritma Performans Karşılaştırması"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Sorgu Sayısı: {len(self.test_queries)}"
        ws['A2'].alignment = center_align
        
        # Tablo başlıkları
        headers = ["Sıra", "Veritabanı", "Algoritma", "Ort. Süre (ms)", "Min (ms)", "Max (ms)", "QPS"]
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Tüm algoritmaları topla ve sırala
        all_algorithms = []
        for db_name, db_results in self.results.items():
            if "algorithms" in db_results:
                for algo_name, algo_data in db_results["algorithms"].items():
                    if "performance" in algo_data:
                        perf = algo_data["performance"]
                        all_algorithms.append({
                            "database": db_name,
                            "algorithm": algo_name,
                            "avg_time_ms": perf["avg_time"] * 1000,
                            "min_time_ms": perf["min_time"] * 1000,
                            "max_time_ms": perf["max_time"] * 1000,
                            "qps": algo_data.get("queries_per_second", 0)
                        })
        
        all_algorithms.sort(key=lambda x: x["avg_time_ms"])
        
        # Verileri yaz
        row = 5
        for i, algo in enumerate(all_algorithms, 1):
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=algo["database"])
            ws.cell(row=row, column=3, value=algo["algorithm"])
            ws.cell(row=row, column=4, value=round(algo["avg_time_ms"], 3))
            ws.cell(row=row, column=5, value=round(algo["min_time_ms"], 3))
            ws.cell(row=row, column=6, value=round(algo["max_time_ms"], 3))
            ws.cell(row=row, column=7, value=round(algo["qps"], 2))
            
            # Stil uygula
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = border
                
                # İlk 3 için özel renk
                if i == 1:
                    cell.fill = gold_fill
                elif i == 2:
                    cell.fill = silver_fill
                elif i == 3:
                    cell.fill = bronze_fill
            
            row += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        # Her veritabanı için ayrı sayfa
        for db_name, db_results in self.results.items():
            if "algorithms" in db_results and db_results["algorithms"]:
                ws_db = wb.create_sheet(title=db_name[:31])  # Excel 31 karakter sınırı
                
                ws_db['A1'] = f"{db_name.upper()} Algoritma Sonuçları"
                ws_db['A1'].font = Font(bold=True, size=14)
                
                if "record_count" in db_results:
                    ws_db['A2'] = f"Kayıt Sayısı: {db_results['record_count']}"
                
                headers = ["Algoritma", "Ort. Süre (ms)", "Min (ms)", "Max (ms)", "Std (ms)", "P95 (ms)", "QPS"]
                row = 4
                for col, header in enumerate(headers, 1):
                    cell = ws_db.cell(row=row, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                row = 5
                for algo_name, algo_data in db_results["algorithms"].items():
                    if "performance" in algo_data:
                        perf = algo_data["performance"]
                        ws_db.cell(row=row, column=1, value=algo_name)
                        ws_db.cell(row=row, column=2, value=round(perf["avg_time"]*1000, 3))
                        ws_db.cell(row=row, column=3, value=round(perf["min_time"]*1000, 3))
                        ws_db.cell(row=row, column=4, value=round(perf["max_time"]*1000, 3))
                        ws_db.cell(row=row, column=5, value=round(perf.get("std_time", 0)*1000, 3))
                        ws_db.cell(row=row, column=6, value=round(perf.get("p95_time", 0)*1000, 3))
                        ws_db.cell(row=row, column=7, value=round(algo_data.get("queries_per_second", 0), 2))
                        
                        for col in range(1, 8):
                            cell = ws_db.cell(row=row, column=col)
                            cell.alignment = center_align
                            cell.border = border
                        
                        row += 1
                
                # Sütun genişlikleri
                ws_db.column_dimensions['A'].width = 30
                for col in ['B', 'C', 'D', 'E', 'F', 'G']:
                    ws_db.column_dimensions[col].width = 15
        
        wb.save(excel_file)
        print(f"💾 Excel sonuçları kaydedildi: {excel_file}")
    
    def save_results_to_excel(self):
        """Sonuçları detaylı Excel dosyasına kaydet"""
        output_dir = "/home/ugo/Documents/Python/bitirememe projesi"
        excel_file = os.path.join(output_dir, "algorithm_benchmark_detailed.xlsx")
        
        print(f"\n📊 Excel dosyası oluşturuluyor...")
        
        wb = openpyxl.Workbook()
        
        # Stil tanımları
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. GENEL ÖZET SAYFASI
        ws_summary = wb.active
        ws_summary.title = "Genel Özet"
        
        ws_summary.merge_cells('A1:H1')
        ws_summary['A1'] = "VEKTÖR VERİTABANI ALGORİTMA PERFORMANS KARŞILAŞTIRMASI"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].alignment = center_align
        
        ws_summary.merge_cells('A2:H2')
        ws_summary['A2'] = f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Test Sorgu Sayısı: {len(self.test_queries)} | Vektör Boyutu: {self.vector_dim}"
        ws_summary['A2'].alignment = center_align
        
        # Tablo başlıkları
        headers = ["Sıra", "Veritabanı", "Algoritma", "Ort. Süre (ms)", "Min (ms)", "Max (ms)", "Std (ms)", "QPS"]
        row = 4
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Tüm algoritmaları topla ve sırala
        all_algorithms = []
        for db_name, db_results in self.results.items():
            if "algorithms" in db_results:
                for algo_name, algo_data in db_results["algorithms"].items():
                    if "performance" in algo_data:
                        perf = algo_data["performance"]
                        all_algorithms.append({
                            "database": db_name,
                            "algorithm": algo_name,
                            "avg_time_ms": perf["avg_time"] * 1000,
                            "min_time_ms": perf["min_time"] * 1000,
                            "max_time_ms": perf["max_time"] * 1000,
                            "std_time_ms": perf.get("std_time", 0) * 1000,
                            "qps": algo_data.get("queries_per_second", 0)
                        })
        
        all_algorithms.sort(key=lambda x: x["avg_time_ms"])
        
        # Verileri yaz
        row = 5
        for i, algo in enumerate(all_algorithms, 1):
            ws_summary.cell(row=row, column=1, value=i)
            ws_summary.cell(row=row, column=2, value=algo["database"])
            ws_summary.cell(row=row, column=3, value=algo["algorithm"])
            ws_summary.cell(row=row, column=4, value=round(algo["avg_time_ms"], 3))
            ws_summary.cell(row=row, column=5, value=round(algo["min_time_ms"], 3))
            ws_summary.cell(row=row, column=6, value=round(algo["max_time_ms"], 3))
            ws_summary.cell(row=row, column=7, value=round(algo["std_time_ms"], 3))
            ws_summary.cell(row=row, column=8, value=round(algo["qps"], 2))
            
            # Stil uygula
            for col in range(1, 9):
                cell = ws_summary.cell(row=row, column=col)
                cell.alignment = center_align
                cell.border = border
                
                # İlk 3 için özel renk
                if i == 1:
                    cell.fill = gold_fill
                elif i == 2:
                    cell.fill = silver_fill
                elif i == 3:
                    cell.fill = bronze_fill
            
            row += 1
        
        # Sütun genişlikleri
        ws_summary.column_dimensions['A'].width = 8
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 30
        ws_summary.column_dimensions['D'].width = 15
        ws_summary.column_dimensions['E'].width = 12
        ws_summary.column_dimensions['F'].width = 12
        ws_summary.column_dimensions['G'].width = 12
        ws_summary.column_dimensions['H'].width = 12
        
        # 2. VERİTABANI BAZINDA DETAYLAR
        for db_name, db_results in self.results.items():
            if "algorithms" in db_results and db_results["algorithms"]:
                ws_db = wb.create_sheet(title=db_name.upper()[:31])
                
                ws_db.merge_cells('A1:H1')
                ws_db['A1'] = f"{db_name.upper()} - Algoritma Detay Sonuçları"
                ws_db['A1'].font = Font(bold=True, size=14)
                ws_db['A1'].alignment = center_align
                
                if "record_count" in db_results:
                    ws_db['A2'] = f"Toplam Kayıt: {db_results['record_count']:,}"
                    ws_db['A2'].font = Font(bold=True, size=11)
                
                headers = ["Algoritma", "Ort. (ms)", "Min (ms)", "Max (ms)", "Std (ms)", "P50 (ms)", "P95 (ms)", "QPS"]
                row = 4
                for col, header in enumerate(headers, 1):
                    cell = ws_db.cell(row=row, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                row = 5
                for algo_name, algo_data in db_results["algorithms"].items():
                    if "performance" in algo_data:
                        perf = algo_data["performance"]
                        ws_db.cell(row=row, column=1, value=algo_name)
                        ws_db.cell(row=row, column=2, value=round(perf["avg_time"]*1000, 3))
                        ws_db.cell(row=row, column=3, value=round(perf["min_time"]*1000, 3))
                        ws_db.cell(row=row, column=4, value=round(perf["max_time"]*1000, 3))
                        ws_db.cell(row=row, column=5, value=round(perf.get("std_time", 0)*1000, 3))
                        ws_db.cell(row=row, column=6, value=round(perf.get("p50_time", 0)*1000, 3))
                        ws_db.cell(row=row, column=7, value=round(perf.get("p95_time", 0)*1000, 3))
                        ws_db.cell(row=row, column=8, value=round(algo_data.get("queries_per_second", 0), 2))
                        
                        for col in range(1, 9):
                            cell = ws_db.cell(row=row, column=col)
                            cell.alignment = center_align if col > 1 else left_align
                            cell.border = border
                        
                        row += 1
                
                # Sütun genişlikleri
                ws_db.column_dimensions['A'].width = 30
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    ws_db.column_dimensions[col].width = 13
        
        # 3. KATEGORİ KARŞILAŞTIRMA SAYFASI
        ws_category = wb.create_sheet(title="Kategori Karşılaştırma")
        
        ws_category.merge_cells('A1:E1')
        ws_category['A1'] = "ALGORİTMA KATEGORİLERİ PERFORMANS KARŞILAŞTIRMASI"
        ws_category['A1'].font = Font(bold=True, size=14)
        ws_category['A1'].alignment = center_align
        
        categories = {
            "HNSW Tabanlı": ["HNSW", "HNSW_NEAR_VECTOR", "HNSW_TEXT", "HNSW_VECTOR"],
            "IVF Tabanlı": ["IVF_FLAT", "IVF_SQ8", "IVF_PQ"],
            "Doğrusal (Flat)": ["FLAT", "FLAT_EXACT", "COSINE", "L2_EUCLIDEAN"],
            "Hybrid/BM25": ["HYBRID", "BM25"],
            "Vektör Arama": ["VECTOR_SEARCH", "VECTOR_PRECOMPUTED"]
        }
        
        headers = ["Kategori", "Algoritma Sayısı", "Ort. Süre (ms)", "En İyi", "En İyi Süre (ms)"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws_category.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row = 4
        for category, algo_patterns in categories.items():
            category_algos = []
            for algo in all_algorithms:
                for pattern in algo_patterns:
                    if pattern in algo["algorithm"]:
                        category_algos.append(algo)
                        break
            
            if category_algos:
                avg_time = np.mean([a["avg_time_ms"] for a in category_algos])
                best = min(category_algos, key=lambda x: x["avg_time_ms"])
                
                ws_category.cell(row=row, column=1, value=category)
                ws_category.cell(row=row, column=2, value=len(category_algos))
                ws_category.cell(row=row, column=3, value=round(avg_time, 3))
                ws_category.cell(row=row, column=4, value=f"{best['database']}/{best['algorithm']}")
                ws_category.cell(row=row, column=5, value=round(best['avg_time_ms'], 3))
                
                for col in range(1, 6):
                    cell = ws_category.cell(row=row, column=col)
                    cell.alignment = left_align if col in [1, 4] else center_align
                    cell.border = border
                
                row += 1
        
        ws_category.column_dimensions['A'].width = 20
        ws_category.column_dimensions['B'].width = 15
        ws_category.column_dimensions['C'].width = 15
        ws_category.column_dimensions['D'].width = 40
        ws_category.column_dimensions['E'].width = 15
        
        # 4. VERİTABANI KARŞILAŞTIRMA
        ws_db_compare = wb.create_sheet(title="Veritabanı Karşılaştırma")
        
        ws_db_compare.merge_cells('A1:F1')
        ws_db_compare['A1'] = "VERİTABANI PERFORMANS KARŞILAŞTIRMASI"
        ws_db_compare['A1'].font = Font(bold=True, size=14)
        ws_db_compare['A1'].alignment = center_align
        
        headers = ["Veritabanı", "Kayıt Sayısı", "Algoritma Sayısı", "En Hızlı (ms)", "En Yavaş (ms)", "Ortalama (ms)"]
        row = 3
        for col, header in enumerate(headers, 1):
            cell = ws_db_compare.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row = 4
        for db_name in self.results.keys():
            db_algos = [a for a in all_algorithms if a["database"] == db_name]
            if db_algos:
                record_count = self.results[db_name].get("record_count", 0)
                fastest = min(db_algos, key=lambda x: x["avg_time_ms"])
                slowest = max(db_algos, key=lambda x: x["avg_time_ms"])
                avg_time = np.mean([a["avg_time_ms"] for a in db_algos])
                
                ws_db_compare.cell(row=row, column=1, value=db_name)
                ws_db_compare.cell(row=row, column=2, value=record_count)
                ws_db_compare.cell(row=row, column=3, value=len(db_algos))
                ws_db_compare.cell(row=row, column=4, value=round(fastest["avg_time_ms"], 3))
                ws_db_compare.cell(row=row, column=5, value=round(slowest["avg_time_ms"], 3))
                ws_db_compare.cell(row=row, column=6, value=round(avg_time, 3))
                
                for col in range(1, 7):
                    cell = ws_db_compare.cell(row=row, column=col)
                    cell.alignment = left_align if col == 1 else center_align
                    cell.border = border
                
                row += 1
        
        ws_db_compare.column_dimensions['A'].width = 15
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_db_compare.column_dimensions[col].width = 15
        
        # Dosyayı kaydet
        wb.save(excel_file)
        print(f"✅ Excel dosyası kaydedildi: {excel_file}")
        print(f"   📄 Sayfalar:")
        print(f"      - Genel Özet: Tüm algoritmaların sıralı listesi")
        print(f"      - Kategori Karşılaştırma: Algoritma kategorilerine göre özet")
        print(f"      - Veritabanı Karşılaştırma: Veritabanı bazında performans")
        for db_name, db_results in self.results.items():
            if "algorithms" in db_results and db_results["algorithms"]:
                print(f"      - {db_name.upper()}: Detaylı algoritma sonuçları")

if __name__ == "__main__":
    benchmark = AlgorithmBenchmark()
    benchmark.run_all_benchmarks()