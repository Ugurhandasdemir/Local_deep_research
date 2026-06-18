"""
Vektor Veritabani Benchmark — Orchestrator + Worker subprocess
----------------------------------------------------------------
Amac: 4GB VRAM + sinirli RAM ortaminda 20+ embedding modelini 5 vektor
veritabaninda guvenli sekilde benchmark etmek.

Tasarim:
  - Her (model, db) kombinasyonu AYRI subprocess'te calisir.
  - Subprocess cikinca tum RAM/VRAM tamamen serbest.
  - WRITE_BATCH=32 — chromadb/lancedb OOM riskini dusurur.
  - Embeddingler bir kez encode edilip embeddings_cache/*.npy'a yazilir;
    sonraki write/search worker'lari .npy'dan okur (yeniden encode YOK).
  - Her test sonucu hemen JSON + Excel'e yazilir.
  - Resume: yarida kalip yeniden baslatildiginda biten testler atlanir.
  - ASAMA 3 — BULMA BASARIMI: SQuAD korpusu uzerinde nDCG@10, Recall@10/100,
    MRR@10, Hit@1/10 (BEIR-style). Yapilan testler JSON kontrolu ile atlanir.

Kullanim:
    python write_vector_database.py --mode test              # 20 belge
    python write_vector_database.py --mode full              # tum belgeler
    python write_vector_database.py --mode test --reset      # cache+sonuc sifirla

(Worker bayragi otomatik, manuel cagirmaya gerek yok)
"""

import os
import gc
import sys
import json
import time
import shutil
import argparse
import warnings
import logging
import threading
import subprocess
from datetime import datetime

import numpy as np


# ---------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------
LOG_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

logger = logging.getLogger("benchmark")


def setup_logging(role):
    log_file = os.path.join(LOG_DIR, f"benchmark_{role}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(sh)
    return log_file


class _PrintToLogger:
    """print() -> logger. Mevcut print cagrilari log dosyasina dussun."""
    def __init__(self, level=logging.INFO):
        self.level = level
        self._buf = ""
    def write(self, msg):
        self._buf += msg
        while "\n" in self._buf:
            line, self._buf = self._buf.split("\n", 1)
            if line.strip():
                logger.log(self.level, line.rstrip())
    def flush(self):
        if self._buf.strip():
            logger.log(self.level, self._buf.rstrip())
        self._buf = ""
    def isatty(self): return False
    def fileno(self): raise OSError("no fileno")
    def writable(self): return True
    def readable(self): return False
    def seekable(self): return False
    @property
    def encoding(self): return "utf-8"
    @property
    def closed(self): return False


def _heartbeat(interval=30):
    try:
        import torch
    except Exception:
        torch = None
    last_wall = time.time()
    last_mono = time.monotonic()
    while True:
        time.sleep(interval)
        now_wall = time.time()
        now_mono = time.monotonic()
        wall_gap = now_wall - last_wall
        mono_gap = now_mono - last_mono
        drift = wall_gap - mono_gap
        try:
            mem_used = torch.cuda.memory_allocated() / 1e9 if (torch and torch.cuda.is_available()) else 0
        except Exception:
            mem_used = 0
        if drift > interval * 1.5:
            logger.warning(
                f"HEARTBEAT — SUSPEND TESPIT EDILDI: {drift:.0f}s atlama "
                f"(wall={wall_gap:.0f}s mono={mono_gap:.0f}s)"
            )
        else:
            logger.info(f"HEARTBEAT alive vram={mem_used:.2f}GB")
        last_wall = now_wall
        last_mono = now_mono


def _inhibit_suspend():
    try:
        p = subprocess.Popen(
            [
                "systemd-inhibit",
                "--what=sleep:idle:handle-lid-switch",
                "--who=vektor-benchmark",
                "--why=Benchmark calisiyor",
                "--mode=block",
                "sleep", "infinity",
            ],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        logger.info(f"systemd-inhibit aktif (pid={p.pid}) — suspend/lid kapali")
        return p
    except FileNotFoundError:
        logger.warning("systemd-inhibit bulunamadi — suspend engellenemedi")
        return None


warnings.filterwarnings("ignore")


# ---------------------------------------------------------------
# AYARLAR
# ---------------------------------------------------------------
BASE_DIR = "/home/ugo/Documents/Python/bitirememe projesi"
MODELS_DIR = os.path.join(BASE_DIR, "models")
DB_DIR = os.path.join(BASE_DIR, "DB")
DATA_FILE = os.path.join(BASE_DIR, "CUSTOM_DATASET", "metin_dosyasi.json")
CACHE_DIR = os.path.join(BASE_DIR, "embeddings_cache")
os.makedirs(CACHE_DIR, exist_ok=True)

MODELS = {
    # --- LOCAL fine-tuned modeller ---
    "e5_small":           {"path": os.path.join(MODELS_DIR, "e5_small", "model"),    "dim": 384},
    "mpnet_multi":        {"path": os.path.join(MODELS_DIR, "mpnet_multi", "model"), "dim": 768},
    "e5_base":            {"path": os.path.join(MODELS_DIR, "e5_base", "model"),     "dim": 768},
    "bge_squad":          {"path": os.path.join(MODELS_DIR, "bge_squad_model"),      "dim": 1024},
    "qwen_lora":          {"path": os.path.join(MODELS_DIR, "qwen_lora"),             "dim": 1024, "batch": 2, "max_seq_len": 128},
    "snowflake_arctic_l": {"path": os.path.join(MODELS_DIR, "snowflake-arctic-embed-l-v2.0"), "dim": 1024},
    "all_mini_l6":        {"path": os.path.join(MODELS_DIR, "all_mini_l6_v2"),         "dim": 384},
    "bge_m3_fine":        {"path": os.path.join(MODELS_DIR, "bge-m3-fine"),         "dim": 1024},


    # --- Base (egitilmemis) HuggingFace versiyonlari ---
    "e5_small_base":          {"path": "intfloat/multilingual-e5-small",                   "dim": 384},
    "mpnet_multi_base":       {"path": "sentence-transformers/paraphrase-multilingual-mpnet-base-v2", "dim": 768},
    "e5_base_base":           {"path": "intfloat/multilingual-e5-base",                    "dim": 768},
    "bge_squad_base":         {"path": "BAAI/bge-large-en-v1.5",                           "dim": 1024},
    "qwen_lora_base":         {"path": "Qwen/Qwen3-Embedding-0.6B",                        "dim": 1024, "batch": 2, "max_seq_len": 128},
    "snowflake_arctic_l_base": {"path": "Snowflake/snowflake-arctic-embed-l-v2.0",         "dim": 1024},
    "all_mini_l6_base":       {"path": "sentence-transformers/all-MiniLM-L6-v2",           "dim": 384},
    "bge_m3_base":            {"path": "BAAI/bge-m3",                                       "dim": 1024},

    # --- Diger HuggingFace hazir modeller ---
    "minilm_l12":         {"path": "sentence-transformers/all-MiniLM-L12-v2",                     "dim": 384},
    "mpnet_base":         {"path": "sentence-transformers/all-mpnet-base-v2",                     "dim": 768},
    "distilroberta":      {"path": "sentence-transformers/all-distilroberta-v1",                  "dim": 768},
    "multi_qa_minilm":    {"path": "sentence-transformers/multi-qa-MiniLM-L6-cos-v1",             "dim": 384},
    "multi_qa_mpnet":     {"path": "sentence-transformers/multi-qa-mpnet-base-dot-v1",            "dim": 768},
    "paraphrase_multi":   {"path": "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2", "dim": 384},
    "bge_small_en":       {"path": "BAAI/bge-small-en-v1.5",                                      "dim": 384},
    "bge_base_en":        {"path": "BAAI/bge-base-en-v1.5",                                       "dim": 768},
    "gte_small":          {"path": "thenlper/gte-small",                                          "dim": 384},
    "gte_base":           {"path": "thenlper/gte-base",                                           "dim": 768},
    "e5_small_hf":        {"path": "intfloat/e5-small-v2",                                        "dim": 384},
    "e5_base_hf":         {"path": "intfloat/e5-base-v2",                                         "dim": 768},
}

DBS = ["milvus", "qdrant", "chromadb", "lancedb", "weaviate"]

# Statik fallback — SQuAD bulunamazsa kullanilir
_FALLBACK_QUERIES = [
    "artificial intelligence healthcare applications",
    "machine learning medical diagnosis systems",
    "deep learning neural network architectures",
    "natural language processing techniques",
    "computer vision medical imaging analysis",
    "reinforcement learning robotics control",
    "transformer models attention mechanism",
    "convolutional neural networks image classification",
    "recurrent neural networks sequence modeling",
    "generative adversarial networks image synthesis",
]

WARMUP_RUNS = 2
TEST_RUNS = 3      # 1000 sorgu x 3 run = 3000 olcum, p99 stabil
ENCODE_BATCH = 32
WRITE_BATCH = 32   # chromadb/lancedb OOM riskini dusurmek icin kucuk

# --- BULMA BASARIMI (akademik / BEIR-style retrieval quality) ---
SQUAD_FILE = os.path.join(BASE_DIR, "CUSTOM_DATASET", "squad_dataset.json")
QUALITY_QUERY_SAMPLE = 1000     # akademik raporlar icin yeterli ornek
QUALITY_TOP_K = 100             # max retrieve, sonra @1/@10/@100 hesapla
QUALITY_SAMPLE_SEED = 42

# --- HIZ BASARIMI (QPS / p50 / p95 / p99) ---
PERF_QUERY_SAMPLE = 1000        # akademik QPS/p99 icin 1000 sorgu
PERF_SAMPLE_SEED = 1337


def load_perf_queries(n=PERF_QUERY_SAMPLE, seed=PERF_SAMPLE_SEED):
    """SQuAD'den N adet sorgu yukle. Bulamazsa fallback listeyi dondur."""
    if not os.path.exists(SQUAD_FILE):
        return list(_FALLBACK_QUERIES)
    try:
        with open(SQUAD_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        qs = []
        for article in data.get("data", []):
            for para in article.get("paragraphs", []):
                for qa in para.get("qas", []):
                    if qa.get("is_impossible", False):
                        continue
                    q = (qa.get("question") or "").strip()
                    if q:
                        qs.append(q)
        if not qs:
            return list(_FALLBACK_QUERIES)
        import random
        rng = random.Random(seed)
        if len(qs) > n:
            qs = rng.sample(qs, n)
        return qs
    except Exception:
        return list(_FALLBACK_QUERIES)


TEST_QUERIES = load_perf_queries()

OUTPUT_JSON = os.path.join(BASE_DIR, "multi_model_benchmark_results.json")
OUTPUT_XLSX = os.path.join(BASE_DIR, "multi_model_benchmark_results.xlsx")


# ---------------------------------------------------------------
# Yardimcilar
# ---------------------------------------------------------------
def free_memory():
    gc.collect()
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            torch.cuda.synchronize()
    except Exception:
        pass


def print_header(title):
    print("\n" + "=" * 70)
    print(" " + title)
    print("=" * 70)


def load_documents(limit=None):
    if not os.path.exists(DATA_FILE):
        raise FileNotFoundError(f"Veri dosyasi bulunamadi: {DATA_FILE}")
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    pdf_docs = data.get("documents", [])
    chunks = []
    chunk_size = 500
    overlap = 50
    step = chunk_size - overlap
    for doc in pdf_docs:
        text = doc.get("full_text", "").strip()
        filename = doc.get("filename", "unknown")
        if not text:
            continue
        for i in range(0, len(text), step):
            chunk = text[i:i + chunk_size].strip()
            if len(chunk) > 50:
                chunks.append({"text": chunk, "source": filename})
    if limit is not None:
        chunks = chunks[:limit]
    return chunks


def measure_time(func):
    for _ in range(WARMUP_RUNS):
        try:
            func()
        except Exception as e:
            return {"error": str(e)}
    times = []
    for _ in range(TEST_RUNS):
        t0 = time.time()
        try:
            func()
        except Exception as e:
            return {"error": str(e)}
        times.append(time.time() - t0)
    return {
        "avg_time": float(np.mean(times)),
        "min_time": float(np.min(times)),
        "max_time": float(np.max(times)),
        "std_time": float(np.std(times)),
        "p50_time": float(np.percentile(times, 50)),
        "p95_time": float(np.percentile(times, 95)),
    }


def cache_paths(model_key, mode):
    return {
        "docs": os.path.join(CACHE_DIR, f"{model_key}_{mode}_docs.npy"),
        "queries": os.path.join(CACHE_DIR, f"{model_key}_{mode}_queries.npy"),
        "meta": os.path.join(CACHE_DIR, f"{model_key}_{mode}_meta.json"),
    }


def encoding_cached(model_key, mode):
    cp = cache_paths(model_key, mode)
    return all(os.path.exists(cp[k]) for k in ("docs", "queries", "meta"))


# ---------------------------------------------------------------
# BULMA BASARIMI — SQuAD korpus + qrels + cache
# ---------------------------------------------------------------
def load_squad_quality():
    """
    SQuAD JSON -> (corpus, queries, qrels)
      corpus : [{"id": int, "text": str, "title": str}]
      queries: [{"id": str, "text": str}]   (deterministik subset)
      qrels  : {qid: [doc_id, ...]}
    """
    if not os.path.exists(SQUAD_FILE):
        raise FileNotFoundError(f"SQuAD bulunamadi: {SQUAD_FILE}")
    with open(SQUAD_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    corpus = []
    ctx_to_id = {}
    all_queries = []
    all_qrels = {}
    for article in data.get("data", []):
        title = article.get("title", "")
        for para in article.get("paragraphs", []):
            ctx = (para.get("context") or "").strip()
            if not ctx:
                continue
            if ctx not in ctx_to_id:
                cid = len(corpus)
                ctx_to_id[ctx] = cid
                corpus.append({"id": cid, "text": ctx, "title": title})
            cid = ctx_to_id[ctx]
            for qa in para.get("qas", []):
                if qa.get("is_impossible", False):
                    continue
                q = (qa.get("question") or "").strip()
                qid = qa.get("id") or ""
                if not q or not qid:
                    continue
                all_queries.append({"id": qid, "text": q})
                all_qrels.setdefault(qid, set()).add(cid)

    import random
    rng = random.Random(QUALITY_SAMPLE_SEED)
    if len(all_queries) > QUALITY_QUERY_SAMPLE:
        all_queries = rng.sample(all_queries, QUALITY_QUERY_SAMPLE)
    qrels = {q["id"]: sorted(all_qrels[q["id"]]) for q in all_queries}
    return corpus, all_queries, qrels


def qual_cache_paths(model_key):
    return {
        "corpus":  os.path.join(CACHE_DIR, f"{model_key}_qual_corpus.npy"),
        "queries": os.path.join(CACHE_DIR, f"{model_key}_qual_queries.npy"),
        "meta":    os.path.join(CACHE_DIR, f"{model_key}_qual_meta.json"),
    }


def qual_encoding_cached(model_key):
    qp = qual_cache_paths(model_key)
    return all(os.path.exists(qp[k]) for k in ("corpus", "queries", "meta"))


def _resolve_encode_device(cfg):
    """Model config 'device' override veya cuda varsa cuda, yoksa cpu."""
    import torch
    forced = cfg.get("device")
    if forced:
        if forced == "cuda" and not torch.cuda.is_available():
            return "cpu"
        return forced
    return "cuda" if torch.cuda.is_available() else "cpu"


def encode_with_fallback(cfg, texts_list, show_progress=True):
    """
    SentenceTransformer encode + OOM fallback.
      - cfg: MODELS[key] (path, dim, optional: device, batch, max_seq_len)
      - texts_list: [list[str], list[str], ...]  birden fazla grup
    Donen: aynı sirada [np.ndarray, ...] + kullanilan device + batch.
    OOM (CUDA) yakalanirsa CPU'ya dusup yeniden dener.
    """
    from sentence_transformers import SentenceTransformer
    import torch

    device = _resolve_encode_device(cfg)
    batch = int(cfg.get("batch", ENCODE_BATCH))
    max_seq = cfg.get("max_seq_len")

    def _load(dev):
        m = SentenceTransformer(cfg["path"], device=dev)
        if max_seq:
            try:
                m.max_seq_length = int(max_seq)
            except Exception:
                pass
        return m

    def _encode_all(m, b):
        out = []
        for i, texts in enumerate(texts_list):
            embs = m.encode(
                texts, batch_size=b, convert_to_numpy=True,
                normalize_embeddings=True,
                show_progress_bar=(show_progress and i == 0),
            )
            out.append(embs)
        return out

    # Cihaz/Bellek konfigurasyonu (OOM riskli modeller icin parcalanma azaltma)
    os.environ.setdefault("PYTORCH_CUDA_ALLOC_CONF", "expandable_segments:True")

    cur_batch = batch
    last_err = None
    out = None
    while cur_batch >= 1:
        try:
            model = _load(device)
            out = _encode_all(model, cur_batch)
            del model
            free_memory()
            return out, device, cur_batch
        except torch.cuda.OutOfMemoryError as e:
            last_err = e
            logger.warning(f"CUDA OOM batch={cur_batch}: {e}. Batch yariya inip tekrar.")
            try:
                del model
            except Exception:
                pass
            free_memory()
            if cur_batch == 1:
                break
            cur_batch = max(1, cur_batch // 2)
        except Exception as e:
            logger.exception(f"Encode hata: {e}")
            free_memory()
            raise
    raise RuntimeError(f"CUDA OOM batch=1'de bile cozulemedi: {last_err}")


def compute_quality_metrics(retrieved_ids_list, gt_list, ks=(1, 10, 100)):
    """
    BEIR-style metrik. Tek qrels (binary relevance) varsayimi.
      retrieved_ids_list: list per-query of top-k doc id (int)
      gt_list           : list per-query of relevant doc id list
    """
    import math
    n = len(retrieved_ids_list)
    if n == 0:
        return {}
    out = {}
    for k in ks:
        hits = 0.0
        recalls = 0.0
        ndcgs = 0.0
        for retrieved, gt in zip(retrieved_ids_list, gt_list):
            top = retrieved[:k]
            gt_set = set(gt)
            if not gt_set:
                continue
            n_rel = len(gt_set)
            if any(d in gt_set for d in top):
                hits += 1.0
            recalls += sum(1 for d in top if d in gt_set) / n_rel
            dcg = sum(1.0 / math.log2(i + 2) for i, d in enumerate(top) if d in gt_set)
            idcg = sum(1.0 / math.log2(i + 2) for i in range(min(k, n_rel)))
            if idcg > 0:
                ndcgs += dcg / idcg
        out[f"hit@{k}"]    = hits / n
        out[f"recall@{k}"] = recalls / n
        out[f"ndcg@{k}"]   = ndcgs / n
    mrr = 0.0
    for retrieved, gt in zip(retrieved_ids_list, gt_list):
        gt_set = set(gt)
        for i, d in enumerate(retrieved[:10]):
            if d in gt_set:
                mrr += 1.0 / (i + 1)
                break
    out["mrr@10"] = mrr / n
    return out


# ---------------------------------------------------------------
# WRITE — her DB icin bir fonksiyon (worker icinde calisir)
# ---------------------------------------------------------------
def write_to_milvus(model_key, docs, embeddings):
    from pymilvus import MilvusClient
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "milvus", f"{model_key}_db.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    if os.path.exists(db_path):
        os.remove(db_path)
    client = MilvusClient(db_path)
    t0 = time.time()
    client.create_collection(
        collection_name=collection,
        dimension=len(embeddings[0]),
        metric_type="COSINE",
    )
    for i in range(0, len(docs), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(docs))
        chunk = [
            {"id": k, "vector": embeddings[k], "text": docs[k]["text"][:500], "source": docs[k]["source"]}
            for k in range(i, j)
        ]
        client.insert(collection_name=collection, data=chunk)
    return time.time() - t0


def write_to_qdrant(model_key, docs, embeddings):
    from qdrant_client import QdrantClient
    from qdrant_client.models import Distance, VectorParams, PointStruct
    client = QdrantClient(host="localhost", port=6333, timeout=120)
    collection = f"docs_{model_key}"
    try:
        client.delete_collection(collection)
    except Exception:
        pass
    t0 = time.time()
    client.create_collection(
        collection_name=collection,
        vectors_config=VectorParams(size=len(embeddings[0]), distance=Distance.COSINE),
    )
    for i in range(0, len(docs), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(docs))
        points = [
            PointStruct(id=k, vector=embeddings[k], payload={"text": docs[k]["text"], "source": docs[k]["source"]})
            for k in range(i, j)
        ]
        client.upsert(collection_name=collection, points=points)
    return time.time() - t0


def write_to_chromadb(model_key, docs, embeddings):
    import chromadb
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "chromadb", f"{model_key}_db")
    if os.path.exists(db_path):
        shutil.rmtree(db_path)
    os.makedirs(db_path, exist_ok=True)
    client = chromadb.PersistentClient(path=db_path)
    t0 = time.time()
    col = client.create_collection(name=collection, metadata={"hnsw:space": "cosine"})
    for i in range(0, len(docs), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(docs))
        col.add(
            ids=[str(k) for k in range(i, j)],
            embeddings=embeddings[i:j],
            documents=[docs[k]["text"] for k in range(i, j)],
            metadatas=[{"source": docs[k]["source"]} for k in range(i, j)],
        )
    return time.time() - t0


def write_to_lancedb(model_key, docs, embeddings):
    import lancedb
    table = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "lancedb", f"{model_key}_db")
    if os.path.exists(db_path):
        shutil.rmtree(db_path)
    os.makedirs(db_path, exist_ok=True)
    db = lancedb.connect(db_path)
    t0 = time.time()
    end_first = min(WRITE_BATCH, len(docs))
    first = [
        {"id": k, "vector": embeddings[k], "text": docs[k]["text"], "source": docs[k]["source"]}
        for k in range(0, end_first)
    ]
    tbl = db.create_table(table, data=first)
    for i in range(WRITE_BATCH, len(docs), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(docs))
        chunk = [
            {"id": k, "vector": embeddings[k], "text": docs[k]["text"], "source": docs[k]["source"]}
            for k in range(i, j)
        ]
        tbl.add(chunk)
    return time.time() - t0


def write_to_weaviate(model_key, docs, embeddings):
    import weaviate
    from weaviate.classes.config import Property, DataType, Configure
    client = weaviate.connect_to_local()
    try:
        collection_name = f"Docs_{model_key}"
        try:
            client.collections.delete(collection_name)
        except Exception:
            pass
        t0 = time.time()
        col = client.collections.create(
            name=collection_name,
            vectorizer_config=Configure.Vectorizer.none(),
            properties=[
                Property(name="text", data_type=DataType.TEXT),
                Property(name="source", data_type=DataType.TEXT),
            ],
        )
        with col.batch.dynamic() as batch:
            for d, emb in zip(docs, embeddings):
                batch.add_object(
                    properties={"text": d["text"], "source": d["source"]},
                    vector=emb,
                )
        return time.time() - t0
    finally:
        client.close()


WRITE_FUNCS = {
    "milvus":   write_to_milvus,
    "qdrant":   write_to_qdrant,
    "chromadb": write_to_chromadb,
    "lancedb":  write_to_lancedb,
    "weaviate": write_to_weaviate,
}


# ---------------------------------------------------------------
# QUALITY WRITE — SQuAD korpusu icin ayri koleksiyon
# ---------------------------------------------------------------
def qual_write_milvus(model_key, corpus, embeddings):
    from pymilvus import MilvusClient
    collection = f"docs_qual_{model_key}"
    db_path = os.path.join(DB_DIR, "milvus", f"{model_key}_qual_db.db")
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    if os.path.exists(db_path):
        os.remove(db_path)
    client = MilvusClient(db_path)
    t0 = time.time()
    client.create_collection(
        collection_name=collection,
        dimension=len(embeddings[0]),
        metric_type="COSINE",
    )
    for i in range(0, len(corpus), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(corpus))
        chunk = [
            {"id": int(corpus[k]["id"]),
             "vector": embeddings[k],
             "text": corpus[k]["text"][:500],
             "title": corpus[k].get("title", "")}
            for k in range(i, j)
        ]
        client.insert(collection_name=collection, data=chunk)
    return time.time() - t0


def qual_write_qdrant(model_key, corpus, embeddings):
    from qdrant_client import QdrantClient
    from qdrant_client.models import Distance, VectorParams, PointStruct
    client = QdrantClient(host="localhost", port=6333, timeout=120)
    collection = f"docs_qual_{model_key}"
    try:
        client.delete_collection(collection)
    except Exception:
        pass
    t0 = time.time()
    client.create_collection(
        collection_name=collection,
        vectors_config=VectorParams(size=len(embeddings[0]), distance=Distance.COSINE),
    )
    for i in range(0, len(corpus), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(corpus))
        points = [
            PointStruct(id=int(corpus[k]["id"]), vector=embeddings[k],
                        payload={"text": corpus[k]["text"], "title": corpus[k].get("title", "")})
            for k in range(i, j)
        ]
        client.upsert(collection_name=collection, points=points)
    return time.time() - t0


def qual_write_chromadb(model_key, corpus, embeddings):
    import chromadb
    collection = f"docs_qual_{model_key}"
    db_path = os.path.join(DB_DIR, "chromadb", f"{model_key}_qual_db")
    if os.path.exists(db_path):
        shutil.rmtree(db_path)
    os.makedirs(db_path, exist_ok=True)
    client = chromadb.PersistentClient(path=db_path)
    t0 = time.time()
    col = client.create_collection(name=collection, metadata={"hnsw:space": "cosine"})
    for i in range(0, len(corpus), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(corpus))
        col.add(
            ids=[str(corpus[k]["id"]) for k in range(i, j)],
            embeddings=embeddings[i:j],
            documents=[corpus[k]["text"] for k in range(i, j)],
            metadatas=[{"title": corpus[k].get("title", "")} for k in range(i, j)],
        )
    return time.time() - t0


def qual_write_lancedb(model_key, corpus, embeddings):
    import lancedb
    table = f"docs_qual_{model_key}"
    db_path = os.path.join(DB_DIR, "lancedb", f"{model_key}_qual_db")
    if os.path.exists(db_path):
        shutil.rmtree(db_path)
    os.makedirs(db_path, exist_ok=True)
    db = lancedb.connect(db_path)
    t0 = time.time()
    end_first = min(WRITE_BATCH, len(corpus))
    first = [
        {"id": int(corpus[k]["id"]), "vector": embeddings[k],
         "text": corpus[k]["text"], "title": corpus[k].get("title", "")}
        for k in range(0, end_first)
    ]
    tbl = db.create_table(table, data=first)
    for i in range(WRITE_BATCH, len(corpus), WRITE_BATCH):
        j = min(i + WRITE_BATCH, len(corpus))
        chunk = [
            {"id": int(corpus[k]["id"]), "vector": embeddings[k],
             "text": corpus[k]["text"], "title": corpus[k].get("title", "")}
            for k in range(i, j)
        ]
        tbl.add(chunk)
    return time.time() - t0


def qual_write_weaviate(model_key, corpus, embeddings):
    import weaviate
    from weaviate.classes.config import Property, DataType, Configure
    client = weaviate.connect_to_local()
    try:
        collection_name = f"DocsQual_{model_key}"
        try:
            client.collections.delete(collection_name)
        except Exception:
            pass
        t0 = time.time()
        col = client.collections.create(
            name=collection_name,
            vectorizer_config=Configure.Vectorizer.none(),
            properties=[
                Property(name="text", data_type=DataType.TEXT),
                Property(name="title", data_type=DataType.TEXT),
                Property(name="doc_id", data_type=DataType.INT),
            ],
        )
        with col.batch.dynamic() as batch:
            for c, emb in zip(corpus, embeddings):
                batch.add_object(
                    properties={"text": c["text"], "title": c.get("title", ""),
                                "doc_id": int(c["id"])},
                    vector=emb,
                )
        return time.time() - t0
    finally:
        client.close()


QUAL_WRITE_FUNCS = {
    "milvus":   qual_write_milvus,
    "qdrant":   qual_write_qdrant,
    "chromadb": qual_write_chromadb,
    "lancedb":  qual_write_lancedb,
    "weaviate": qual_write_weaviate,
}


# ---------------------------------------------------------------
# SEARCH — her DB icin bir fonksiyon
# ---------------------------------------------------------------
def search_milvus(model_key, query_vectors):
    from pymilvus import MilvusClient
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "milvus", f"{model_key}_db.db")
    client = MilvusClient(db_path)
    results = {}

    def default_search():
        return [client.search(collection_name=collection, data=[q], limit=10, output_fields=["text"])
                for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r: results["HNSW_default"] = r

    for limit in [5, 20, 50]:
        def fn(lim=limit):
            return [client.search(collection_name=collection, data=[q], limit=lim, output_fields=["text"])
                    for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r: results[f"HNSW_limit{limit}"] = r

    def batch():
        return client.search(collection_name=collection, data=query_vectors, limit=10, output_fields=["text"])
    r = measure_time(batch)
    if "error" not in r: results["HNSW_batch"] = r
    return results


def search_qdrant(model_key, query_vectors):
    from qdrant_client import QdrantClient
    from qdrant_client.models import SearchParams
    client = QdrantClient(host="localhost", port=6333, timeout=60)
    collection = f"docs_{model_key}"
    results = {}

    def default_search():
        return [client.query_points(collection_name=collection, query=q, limit=10) for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r: results["HNSW_default"] = r

    def exact():
        return [client.query_points(collection_name=collection, query=q, limit=10,
                                    search_params=SearchParams(exact=True))
                for q in query_vectors]
    r = measure_time(exact)
    if "error" not in r: results["EXACT"] = r

    for ef in [32, 128, 256]:
        def fn(e=ef):
            return [client.query_points(collection_name=collection, query=q, limit=10,
                                        search_params=SearchParams(hnsw_ef=e))
                    for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r: results[f"HNSW_ef{ef}"] = r
    return results


def search_chromadb(model_key, query_vectors):
    import chromadb
    collection = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "chromadb", f"{model_key}_db")
    client = chromadb.PersistentClient(path=db_path)
    col = client.get_collection(collection)
    results = {}

    def default_search():
        return [col.query(query_embeddings=[q], n_results=10) for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r: results["HNSW_default"] = r

    for n in [5, 20, 50]:
        def fn(nn=n):
            return [col.query(query_embeddings=[q], n_results=nn) for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r: results[f"HNSW_n{n}"] = r

    def batch():
        return col.query(query_embeddings=query_vectors, n_results=10)
    r = measure_time(batch)
    if "error" not in r: results["HNSW_batch"] = r
    return results


def search_lancedb(model_key, query_vectors):
    import lancedb
    table_name = f"docs_{model_key}"
    db_path = os.path.join(DB_DIR, "lancedb", f"{model_key}_db")
    db = lancedb.connect(db_path)
    table = db.open_table(table_name)
    results = {}

    def default_search():
        return [table.search(q).limit(10).to_pandas() for q in query_vectors]
    r = measure_time(default_search)
    if "error" not in r: results["VECTOR_default"] = r

    for limit in [5, 20, 50]:
        def fn(lim=limit):
            return [table.search(q).limit(lim).to_pandas() for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r: results[f"VECTOR_limit{limit}"] = r

    for metric in ["cosine", "L2"]:
        def fn(m=metric):
            return [table.search(q).metric(m).limit(10).to_pandas() for q in query_vectors]
        r = measure_time(fn)
        if "error" not in r: results[f"VECTOR_{metric}"] = r
    return results


def search_weaviate(model_key, query_vectors):
    import weaviate
    client = weaviate.connect_to_local()
    try:
        col = client.collections.get(f"Docs_{model_key}")
        results = {}

        def default_search():
            return [col.query.near_vector(near_vector=q, limit=10) for q in query_vectors]
        r = measure_time(default_search)
        if "error" not in r: results["HNSW_default"] = r

        for limit in [5, 20, 50]:
            def fn(lim=limit):
                return [col.query.near_vector(near_vector=q, limit=lim) for q in query_vectors]
            r = measure_time(fn)
            if "error" not in r: results[f"HNSW_limit{limit}"] = r

        def bm25():
            return [col.query.bm25(query=qt, limit=10) for qt in TEST_QUERIES]
        r = measure_time(bm25)
        if "error" not in r: results["BM25"] = r

        for alpha in [0.25, 0.5, 0.75]:
            def fn(a=alpha):
                return [col.query.hybrid(query=qt, vector=q, limit=10, alpha=a)
                        for qt, q in zip(TEST_QUERIES, query_vectors)]
            r = measure_time(fn)
            if "error" not in r: results[f"HYBRID_alpha{alpha}"] = r
        return results
    finally:
        client.close()


SEARCH_FUNCS = {
    "milvus":   search_milvus,
    "qdrant":   search_qdrant,
    "chromadb": search_chromadb,
    "lancedb":  search_lancedb,
    "weaviate": search_weaviate,
}


# ---------------------------------------------------------------
# QUALITY SEARCH — top-K dondur (id), metrik orchestrator'da hesaplanir
# ---------------------------------------------------------------
def qual_search_milvus(model_key, query_vectors, top_k=QUALITY_TOP_K):
    from pymilvus import MilvusClient
    collection = f"docs_qual_{model_key}"
    db_path = os.path.join(DB_DIR, "milvus", f"{model_key}_qual_db.db")
    client = MilvusClient(db_path)
    out = []
    for q in query_vectors:
        res = client.search(collection_name=collection, data=[q], limit=top_k,
                            output_fields=["title"])
        ids = [int(hit["id"]) for hit in res[0]]
        out.append(ids)
    return out


def qual_search_qdrant(model_key, query_vectors, top_k=QUALITY_TOP_K):
    from qdrant_client import QdrantClient
    client = QdrantClient(host="localhost", port=6333, timeout=120)
    collection = f"docs_qual_{model_key}"
    out = []
    for q in query_vectors:
        res = client.query_points(collection_name=collection, query=q, limit=top_k)
        out.append([int(p.id) for p in res.points])
    return out


def qual_search_chromadb(model_key, query_vectors, top_k=QUALITY_TOP_K):
    import chromadb
    collection = f"docs_qual_{model_key}"
    db_path = os.path.join(DB_DIR, "chromadb", f"{model_key}_qual_db")
    client = chromadb.PersistentClient(path=db_path)
    col = client.get_collection(collection)
    out = []
    for q in query_vectors:
        res = col.query(query_embeddings=[q], n_results=top_k)
        ids = [int(x) for x in res["ids"][0]]
        out.append(ids)
    return out


def qual_search_lancedb(model_key, query_vectors, top_k=QUALITY_TOP_K):
    import lancedb
    table_name = f"docs_qual_{model_key}"
    db_path = os.path.join(DB_DIR, "lancedb", f"{model_key}_qual_db")
    db = lancedb.connect(db_path)
    table = db.open_table(table_name)
    out = []
    for q in query_vectors:
        df = table.search(q).limit(top_k).to_pandas()
        out.append([int(x) for x in df["id"].tolist()])
    return out


def qual_search_weaviate(model_key, query_vectors, top_k=QUALITY_TOP_K):
    import weaviate
    client = weaviate.connect_to_local()
    try:
        col = client.collections.get(f"DocsQual_{model_key}")
        out = []
        for q in query_vectors:
            res = col.query.near_vector(near_vector=q, limit=top_k,
                                        return_properties=["doc_id"])
            ids = [int(o.properties["doc_id"]) for o in res.objects]
            out.append(ids)
        return out
    finally:
        client.close()


QUAL_SEARCH_FUNCS = {
    "milvus":   qual_search_milvus,
    "qdrant":   qual_search_qdrant,
    "chromadb": qual_search_chromadb,
    "lancedb":  qual_search_lancedb,
    "weaviate": qual_search_weaviate,
}


# ---------------------------------------------------------------
# WORKER (subprocess olarak calisir)
# ---------------------------------------------------------------
def worker_main(args):
    role = f"worker_{args.phase}_{args.model}_{args.db or 'NA'}"
    setup_logging(role)
    sys.stdout = _PrintToLogger(logging.INFO)
    sys.stderr = _PrintToLogger(logging.ERROR)
    threading.Thread(target=_heartbeat, daemon=True).start()
    logger.info(f"Worker basladi PID={os.getpid()} phase={args.phase} model={args.model} db={args.db} mode={args.mode}")

    # Bellek takibi (RAM rss + CUDA peak)
    try:
        import psutil
        _proc = psutil.Process(os.getpid())
    except Exception:
        _proc = None
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.reset_peak_memory_stats()
    except Exception:
        torch = None

    def _peak_mem():
        ram_mb = 0.0
        vram_gb = 0.0
        try:
            if _proc is not None:
                ram_mb = _proc.memory_info().rss / (1024 * 1024)
        except Exception:
            pass
        try:
            if torch is not None and torch.cuda.is_available():
                vram_gb = torch.cuda.max_memory_allocated() / 1e9
        except Exception:
            pass
        return ram_mb, vram_gb

    cp = cache_paths(args.model, args.mode)
    result = {"status": "error", "error": "unknown"}
    try:
        if args.phase == "encode":
            cfg = MODELS[args.model]
            logger.info(f"Encode: {args.model} path={cfg['path']} "
                        f"device_override={cfg.get('device')} batch={cfg.get('batch', ENCODE_BATCH)}")
            docs = load_documents(limit=20 if args.mode == "test" else None)
            texts = [d["text"] for d in docs]
            (doc_embs, query_embs), used_dev, used_batch = encode_with_fallback(
                cfg, [texts, TEST_QUERIES], show_progress=True
            )
            logger.info(f"Encode bitti device={used_dev} batch={used_batch}")
            np.save(cp["docs"], doc_embs)
            np.save(cp["queries"], query_embs)
            with open(cp["meta"], "w", encoding="utf-8") as f:
                json.dump({
                    "docs": docs,
                    "dim": int(doc_embs.shape[1]),
                    "doc_count": len(docs),
                    "model": args.model,
                    "mode": args.mode,
                }, f, ensure_ascii=False)
            result = {"status": "ok", "dim": int(doc_embs.shape[1]), "doc_count": len(docs)}

        elif args.phase == "write":
            embeddings = np.load(cp["docs"]).tolist()
            with open(cp["meta"], "r", encoding="utf-8") as f:
                meta = json.load(f)
            docs = meta["docs"]
            fn = WRITE_FUNCS[args.db]
            logger.info(f"Write -> {args.db} ({len(docs)} kayit)")
            t = fn(args.model, docs, embeddings)
            result = {"status": "ok", "time": t, "record_count": len(docs), "dim": meta["dim"]}

        elif args.phase == "search":
            query_vectors = np.load(cp["queries"]).tolist()
            fn = SEARCH_FUNCS[args.db]
            logger.info(f"Search -> {args.db}")
            algos = fn(args.model, query_vectors)
            result = {"status": "ok", "algorithms": algos}

        elif args.phase == "encode_qual":
            cfg = MODELS[args.model]
            logger.info(f"EncodeQual: {args.model} "
                        f"device_override={cfg.get('device')} batch={cfg.get('batch', ENCODE_BATCH)}")
            corpus, queries, qrels = load_squad_quality()
            corpus_texts = [c["text"] for c in corpus]
            query_texts = [q["text"] for q in queries]
            (corpus_embs, query_embs), used_dev, used_batch = encode_with_fallback(
                cfg, [corpus_texts, query_texts], show_progress=True
            )
            logger.info(f"EncodeQual bitti device={used_dev} batch={used_batch}")
            qp = qual_cache_paths(args.model)
            np.save(qp["corpus"], corpus_embs)
            np.save(qp["queries"], query_embs)
            with open(qp["meta"], "w", encoding="utf-8") as f:
                json.dump({
                    "corpus": corpus,
                    "queries": queries,
                    "qrels": qrels,
                    "dim": int(corpus_embs.shape[1]),
                    "model": args.model,
                }, f, ensure_ascii=False)
            result = {"status": "ok", "dim": int(corpus_embs.shape[1]),
                      "corpus_count": len(corpus), "query_count": len(queries)}

        elif args.phase == "write_qual":
            qp = qual_cache_paths(args.model)
            embeddings = np.load(qp["corpus"]).tolist()
            with open(qp["meta"], "r", encoding="utf-8") as f:
                meta = json.load(f)
            corpus = meta["corpus"]
            fn = QUAL_WRITE_FUNCS[args.db]
            logger.info(f"WriteQual -> {args.db} ({len(corpus)} korpus)")
            t = fn(args.model, corpus, embeddings)
            result = {"status": "ok", "time": t, "corpus_count": len(corpus),
                      "dim": meta["dim"]}

        elif args.phase == "search_qual":
            qp = qual_cache_paths(args.model)
            query_vectors = np.load(qp["queries"]).tolist()
            with open(qp["meta"], "r", encoding="utf-8") as f:
                meta = json.load(f)
            queries = meta["queries"]
            qrels = meta["qrels"]
            fn = QUAL_SEARCH_FUNCS[args.db]
            logger.info(f"SearchQual -> {args.db} (top_k={QUALITY_TOP_K})")
            t0 = time.time()
            retrieved = fn(args.model, query_vectors, top_k=QUALITY_TOP_K)
            elapsed = time.time() - t0
            gt_list = [qrels[q["id"]] for q in queries]
            metrics = compute_quality_metrics(retrieved, gt_list, ks=(1, 10, 100))
            metrics["search_time_sec"] = elapsed
            metrics["query_count"] = len(queries)
            metrics["corpus_count"] = len(meta["corpus"])
            metrics["top_k"] = QUALITY_TOP_K
            result = {"status": "ok", "metrics": metrics}

        else:
            result = {"status": "error", "error": f"unknown phase {args.phase}"}
    except Exception as e:
        logger.exception(f"Worker hata: {e}")
        result = {"status": "error", "error": f"{type(e).__name__}: {e}"}

    # Bellek peak'i her durumda result'a ekle
    ram_mb, vram_gb = _peak_mem()
    if isinstance(result, dict):
        result["peak_ram_mb"] = round(ram_mb, 1)
        result["peak_vram_gb"] = round(vram_gb, 3)

    if args.result_file:
        with open(args.result_file, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, default=str)
    logger.info(
        f"Worker bitti status={result.get('status')} "
        f"peak_ram={ram_mb:.0f}MB peak_vram={vram_gb:.2f}GB"
    )


# ---------------------------------------------------------------
# ORCHESTRATOR
# ---------------------------------------------------------------
def load_results():
    if not os.path.exists(OUTPUT_JSON):
        return None
    try:
        with open(OUTPUT_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def init_results(mode, doc_count):
    return {
        "metadata": {
            "date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "mode": mode,
            "doc_count": doc_count,
            "query_count": len(TEST_QUERIES),
            "quality_query_sample": QUALITY_QUERY_SAMPLE,
            "quality_top_k": QUALITY_TOP_K,
            "models": {k: v["path"] for k, v in MODELS.items()},
        },
        "write": {db: {} for db in DBS},
        "search": {db: {} for db in DBS},
        "quality_write": {db: {} for db in DBS},
        "quality_search": {db: {} for db in DBS},
    }


def is_write_done(results, db, model):
    return results.get("write", {}).get(db, {}).get(model, {}).get("status") == "ok"


def is_search_done(results, db, model):
    s = results.get("search", {}).get(db, {}).get(model)
    if not isinstance(s, dict) or not s:
        return False
    if "error" in s and not any(
        isinstance(v, dict) and "avg_time" in v
        for k, v in s.items() if not str(k).startswith("__")
    ):
        return False
    return any(
        isinstance(v, dict) and "avg_time" in v
        for k, v in s.items() if not str(k).startswith("__")
    )


def is_qual_write_done(results, db, model):
    return results.get("quality_write", {}).get(db, {}).get(model, {}).get("status") == "ok"


def is_qual_search_done(results, db, model):
    q = results.get("quality_search", {}).get(db, {}).get(model)
    if not isinstance(q, dict) or not q:
        return False
    return q.get("status") == "ok" and "metrics" in q


def save_json(results):
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False, default=str)


def spawn_worker(phase, model, mode, db=None):
    result_file = os.path.join(BASE_DIR, f".worker_result_{os.getpid()}.json")
    if os.path.exists(result_file):
        os.remove(result_file)
    cmd = [sys.executable, os.path.abspath(__file__), "--worker",
           "--phase", phase, "--model", model, "--mode", mode,
           "--result-file", result_file]
    if db:
        cmd += ["--db", db]
    logger.info(f"Subprocess: phase={phase} model={model} db={db or '-'}")
    proc = subprocess.run(cmd)
    if not os.path.exists(result_file):
        return {"status": "error", "error": f"worker exit={proc.returncode} no result file (muhtemelen OOM kill)"}
    with open(result_file, "r", encoding="utf-8") as f:
        res = json.load(f)
    try:
        os.remove(result_file)
    except Exception:
        pass
    if proc.returncode != 0 and res.get("status") == "ok":
        res["warning"] = f"subprocess returncode={proc.returncode}"
    return res


def update_excel(results):
    try:
        build_excel(results)
        logger.info(f"Excel guncellendi: {OUTPUT_XLSX}")
    except Exception as e:
        logger.error(f"Excel hatasi: {e}")


def orchestrate(mode, reset=False):
    if reset:
        if os.path.exists(OUTPUT_JSON):
            os.remove(OUTPUT_JSON)
        if os.path.exists(CACHE_DIR):
            shutil.rmtree(CACHE_DIR)
            os.makedirs(CACHE_DIR, exist_ok=True)
        logger.info("Reset: sonuc + embeddings_cache silindi")

    docs = load_documents(limit=20 if mode == "test" else None)
    doc_count = len(docs)

    print_header("SISTEM BILGISI")
    try:
        import torch
        if torch.cuda.is_available():
            gpu = torch.cuda.get_device_name(0)
            vram = torch.cuda.get_device_properties(0).total_memory / 1e9
            print(f"  GPU: {gpu}  ({vram:.2f} GB VRAM)")
        else:
            print("  GPU yok — CPU kullanilacak")
    except Exception:
        print("  Torch import hatasi")
    print(f"  Mode={mode} doc_count={doc_count}")

    results = load_results()
    if results is None:
        results = init_results(mode, doc_count)
    else:
        prev_mode = results.get("metadata", {}).get("mode")
        if prev_mode != mode:
            logger.error(f"Mevcut sonuclar mode={prev_mode}, istenen mode={mode}. --reset ile sifirlayin.")
            sys.exit(2)
        if results.get("metadata", {}).get("doc_count") != doc_count:
            logger.warning(f"doc_count degisti {results['metadata'].get('doc_count')} -> {doc_count}")
            results["metadata"]["doc_count"] = doc_count
        # MODELS listesine yeni model eklenmis olabilir — dict'leri tamamla
        for db in DBS:
            results.setdefault("write", {}).setdefault(db, {})
            results.setdefault("search", {}).setdefault(db, {})
            results.setdefault("quality_write", {}).setdefault(db, {})
            results.setdefault("quality_search", {}).setdefault(db, {})
        results.setdefault("metadata", {}).setdefault("quality_query_sample", QUALITY_QUERY_SAMPLE)
        results["metadata"].setdefault("quality_top_k", QUALITY_TOP_K)
    save_json(results)

    # ASAMA 1: encode (cache yoksa) + write (her DB)
    print_header("ASAMA 1: ENCODE + WRITE")
    encode_peaks = results.setdefault("encode_peaks", {})        # {model: {ram_mb, vram_gb}}
    encode_qual_peaks = results.setdefault("encode_qual_peaks", {})
    for model_key, cfg in MODELS.items():
        pending_writes = [db for db in DBS if not is_write_done(results, db, model_key)]
        if not pending_writes:
            logger.info(f"[{model_key}] tum write tamam, atlaniyor")
            continue

        if not encoding_cached(model_key, mode):
            r = spawn_worker("encode", model_key, mode)
            if r.get("status") != "ok":
                logger.error(f"[{model_key}] encode HATA: {r.get('error')}")
                for db in pending_writes:
                    results["write"][db][model_key] = {"status": "error", "error": f"encode: {r.get('error')}"}
                save_json(results)
                update_excel(results)
                continue
            encode_peaks[model_key] = {
                "peak_ram_mb": r.get("peak_ram_mb", 0.0),
                "peak_vram_gb": r.get("peak_vram_gb", 0.0),
            }
        else:
            logger.info(f"[{model_key}] embedding cache hit")

        for db in pending_writes:
            r = spawn_worker("write", model_key, mode, db=db)
            if r.get("status") == "ok":
                results["write"][db][model_key] = {
                    "status": "ok",
                    "time": r["time"],
                    "record_count": r["record_count"],
                    "dim": r["dim"],
                    "peak_ram_mb": r.get("peak_ram_mb", 0.0),
                    "peak_vram_gb": r.get("peak_vram_gb", 0.0),
                }
                logger.info(
                    f"[{model_key}/{db}] write OK {r['time']:.2f}s "
                    f"ram={r.get('peak_ram_mb', 0):.0f}MB vram={r.get('peak_vram_gb', 0):.2f}GB"
                )
            else:
                results["write"][db][model_key] = {"status": "error", "error": r.get("error", "unknown")}
                logger.error(f"[{model_key}/{db}] write HATA: {r.get('error')}")
            save_json(results)
            update_excel(results)

    # ASAMA 2: search (her DB her model)
    print_header("ASAMA 2: SEARCH")
    for db in DBS:
        for model_key in MODELS:
            if is_search_done(results, db, model_key):
                continue
            if not is_write_done(results, db, model_key):
                continue
            if not encoding_cached(model_key, mode):
                logger.warning(f"[{model_key}/{db}] search: query embedding yok, atlaniyor")
                continue

            r = spawn_worker("search", model_key, mode, db=db)
            if r.get("status") == "ok":
                bucket = dict(r["algorithms"])
                bucket["__peak_ram_mb"] = r.get("peak_ram_mb", 0.0)
                bucket["__peak_vram_gb"] = r.get("peak_vram_gb", 0.0)
                results["search"][db][model_key] = bucket
                for algo, perf in r["algorithms"].items():
                    if isinstance(perf, dict) and "avg_time" in perf:
                        logger.info(f"[{model_key}/{db}] {algo:<22} {perf['avg_time']*1000:8.3f} ms")
                logger.info(
                    f"[{model_key}/{db}] search ram={r.get('peak_ram_mb', 0):.0f}MB "
                    f"vram={r.get('peak_vram_gb', 0):.2f}GB"
                )
            else:
                results["search"][db][model_key] = {"error": r.get("error", "unknown")}
                logger.error(f"[{model_key}/{db}] search HATA: {r.get('error')}")
            save_json(results)
            update_excel(results)

    # ASAMA 3: BULMA BASARIMI (akademik retrieval quality — SQuAD)
    print_header("ASAMA 3: BULMA BASARIMI (SQuAD)")
    for model_key, cfg in MODELS.items():
        pending_qwrites = [db for db in DBS if not is_qual_write_done(results, db, model_key)]
        pending_qsearch = [db for db in DBS if not is_qual_search_done(results, db, model_key)]
        if not pending_qwrites and not pending_qsearch:
            logger.info(f"[{model_key}] quality tum DB tamam, atlaniyor")
            continue

        # encode_qual cache yoksa once encode et
        if (pending_qwrites or pending_qsearch) and not qual_encoding_cached(model_key):
            r = spawn_worker("encode_qual", model_key, mode)
            if r.get("status") != "ok":
                logger.error(f"[{model_key}] encode_qual HATA: {r.get('error')}")
                for db in pending_qwrites:
                    results["quality_write"][db][model_key] = {"status": "error",
                                                                "error": f"encode_qual: {r.get('error')}"}
                save_json(results)
                update_excel(results)
                continue
            encode_qual_peaks[model_key] = {
                "peak_ram_mb": r.get("peak_ram_mb", 0.0),
                "peak_vram_gb": r.get("peak_vram_gb", 0.0),
            }
            logger.info(f"[{model_key}] encode_qual OK corpus={r.get('corpus_count')} q={r.get('query_count')}")
        else:
            if pending_qwrites or pending_qsearch:
                logger.info(f"[{model_key}] quality embedding cache hit")

        # write_qual
        for db in pending_qwrites:
            r = spawn_worker("write_qual", model_key, mode, db=db)
            if r.get("status") == "ok":
                results["quality_write"][db][model_key] = {
                    "status": "ok",
                    "time": r["time"],
                    "corpus_count": r["corpus_count"],
                    "dim": r["dim"],
                    "peak_ram_mb": r.get("peak_ram_mb", 0.0),
                    "peak_vram_gb": r.get("peak_vram_gb", 0.0),
                }
                logger.info(f"[{model_key}/{db}] quality_write OK {r['time']:.2f}s")
            else:
                results["quality_write"][db][model_key] = {"status": "error",
                                                            "error": r.get("error", "unknown")}
                logger.error(f"[{model_key}/{db}] quality_write HATA: {r.get('error')}")
            save_json(results)
            update_excel(results)

        # search_qual (yalnizca write basarili olanlarda)
        for db in DBS:
            if is_qual_search_done(results, db, model_key):
                continue
            if not is_qual_write_done(results, db, model_key):
                continue
            r = spawn_worker("search_qual", model_key, mode, db=db)
            if r.get("status") == "ok":
                results["quality_search"][db][model_key] = {
                    "status": "ok",
                    "metrics": r["metrics"],
                    "peak_ram_mb": r.get("peak_ram_mb", 0.0),
                    "peak_vram_gb": r.get("peak_vram_gb", 0.0),
                }
                m = r["metrics"]
                logger.info(
                    f"[{model_key}/{db}] quality "
                    f"ndcg@10={m.get('ndcg@10', 0):.4f} "
                    f"recall@10={m.get('recall@10', 0):.4f} "
                    f"recall@100={m.get('recall@100', 0):.4f} "
                    f"mrr@10={m.get('mrr@10', 0):.4f}"
                )
            else:
                results["quality_search"][db][model_key] = {"status": "error",
                                                             "error": r.get("error", "unknown")}
                logger.error(f"[{model_key}/{db}] quality_search HATA: {r.get('error')}")
            save_json(results)
            update_excel(results)

    update_excel(results)
    print_summary(results)


def print_summary(results):
    print_header("OZET")
    print(f"  Mod: {results['metadata']['mode']}")
    print(f"  Belge: {results['metadata']['doc_count']}")
    print(f"  Sorgu: {len(TEST_QUERIES)}")

    rows = []
    for db, db_data in results["search"].items():
        for model_key, algos in db_data.items():
            if not isinstance(algos, dict): continue
            for algo, perf in algos.items():
                if isinstance(perf, dict) and "avg_time" in perf:
                    rows.append((db, model_key, algo, perf["avg_time"] * 1000))
    rows.sort(key=lambda x: x[3])

    print("\n  EN HIZLI 10 ARAMA:")
    print("  " + "-" * 70)
    for i, (db, model, algo, ms) in enumerate(rows[:10], 1):
        print(f"  {i:2}. {db:<10} | {model:<14} | {algo:<22} | {ms:8.3f} ms")


# ---------------------------------------------------------------
# Excel
# ---------------------------------------------------------------
def build_excel(results):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    gold = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 1) OZET
    ws = wb.active
    ws.title = "Ozet"
    ws["A1"] = "BENCHMARK OZETI"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Tarih:"
    ws["B3"] = results["metadata"]["date"]
    ws["A4"] = "Mod:"
    ws["B4"] = results["metadata"]["mode"]
    ws["A5"] = "Belge sayisi:"
    ws["B5"] = results["metadata"]["doc_count"]
    ws["A6"] = "Sorgu sayisi:"
    ws["B6"] = len(TEST_QUERIES)

    ws["A8"] = "MODELLER:"
    ws["A8"].font = Font(bold=True)
    row = 9
    for key, cfg in MODELS.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = cfg["path"]
        ws[f"C{row}"] = f"dim={cfg['dim']}"
        row += 1
    for c in ["A", "B", "C"]:
        ws.column_dimensions[c].width = 40

    # 2) YAZMA SURELERI
    ws = wb.create_sheet("Yazma Sureleri")
    ws["A1"] = "MODEL x VERITABANI YAZMA SURELERI (saniye)"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Model"] + DBS
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    row = 4
    for model_key in MODELS:
        ws.cell(row=row, column=1, value=model_key).border = border
        for j, db in enumerate(DBS, 2):
            write_data = results.get("write", {}).get(db, {}).get(model_key, {})
            cell = ws.cell(row=row, column=j)
            if write_data.get("status") == "ok" and isinstance(write_data.get("time"), (int, float)):
                cell.value = round(write_data["time"], 3)
            elif write_data.get("status") == "error":
                cell.value = "HATA"
            else:
                cell.value = "-"
            cell.border = border
            cell.alignment = center
        row += 1

    for c in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 16

    # 3) TUM ARAMA SONUCLARI
    ws = wb.create_sheet("Arama Sonuclari")
    ws["A1"] = "TUM ARAMA TESTLERI (ms)"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Sira", "Veritabani", "Model", "Algoritma", "Ort", "Min", "Max", "Std", "P50", "P95"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    all_rows = []
    for db, db_data in results.get("search", {}).items():
        for model_key, algos in db_data.items():
            if not isinstance(algos, dict):
                continue
            for algo, perf in algos.items():
                if not isinstance(perf, dict) or "avg_time" not in perf:
                    continue
                all_rows.append({
                    "db": db, "model": model_key, "algo": algo,
                    "avg": perf["avg_time"] * 1000,
                    "min": perf["min_time"] * 1000,
                    "max": perf["max_time"] * 1000,
                    "std": perf["std_time"] * 1000,
                    "p50": perf["p50_time"] * 1000,
                    "p95": perf["p95_time"] * 1000,
                })
    all_rows.sort(key=lambda x: x["avg"])

    row = 4
    for i, r in enumerate(all_rows, 1):
        vals = [i, r["db"], r["model"], r["algo"],
                round(r["avg"], 3), round(r["min"], 3), round(r["max"], 3),
                round(r["std"], 3), round(r["p50"], 3), round(r["p95"], 3)]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.alignment = center
            cell.border = border
            if i <= 3:
                cell.fill = gold
        row += 1

    ws.column_dimensions["A"].width = 6
    for c in ["B", "C", "D"]:
        ws.column_dimensions[c].width = 18
    for c in ["E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[c].width = 10

    # 4) Her DB icin ayri sayfa
    for db in DBS:
        ws = wb.create_sheet(f"{db.upper()}")
        ws["A1"] = f"{db.upper()} — Model x Algoritma (ms ortalama)"
        ws["A1"].font = Font(bold=True, size=14)

        algo_set = set()
        for model_data in results.get("search", {}).get(db, {}).values():
            if isinstance(model_data, dict):
                algo_set.update(k for k, v in model_data.items() if isinstance(v, dict) and "avg_time" in v)
        algos = sorted(algo_set)

        headers = ["Model"] + algos
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center

        row = 4
        for model_key in MODELS:
            ws.cell(row=row, column=1, value=model_key)
            m = results.get("search", {}).get(db, {}).get(model_key, {})
            for j, algo in enumerate(algos, 2):
                perf = m.get(algo) if isinstance(m, dict) else None
                cell = ws.cell(row=row, column=j)
                if isinstance(perf, dict) and "avg_time" in perf:
                    cell.value = round(perf["avg_time"] * 1000, 3)
                else:
                    cell.value = "-"
                cell.alignment = center
            row += 1

        ws.column_dimensions["A"].width = 16
        for col_idx in range(2, len(algos) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    # 5) BULMA BASARIMI (akademik retrieval quality — SQuAD)
    metric_cols = ["ndcg@10", "recall@10", "recall@100", "mrr@10", "hit@1", "hit@10"]

    ws = wb.create_sheet("Bulma Basarimi")
    ws["A1"] = "BELGE BULMA BASARIMI — SQuAD (akademik / BEIR-style)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Sorgu ornek: {results['metadata'].get('quality_query_sample', QUALITY_QUERY_SAMPLE)}  "
                f"top_k: {results['metadata'].get('quality_top_k', QUALITY_TOP_K)}  "
                "metrik: nDCG@10 (birincil), Recall@10/100, MRR@10, Hit@1/10")

    headers = ["Sira", "Veritabani", "Model"] + metric_cols + ["search_time_sec"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    quality_rows = []
    for db, db_data in results.get("quality_search", {}).items():
        for model_key, payload in db_data.items():
            if not isinstance(payload, dict) or payload.get("status") != "ok":
                continue
            m = payload.get("metrics", {})
            quality_rows.append({
                "db": db, "model": model_key,
                "metrics": m,
                "search_time": m.get("search_time_sec", 0.0),
            })
    quality_rows.sort(key=lambda x: x["metrics"].get("ndcg@10", 0.0), reverse=True)

    row = 5
    for i, r in enumerate(quality_rows, 1):
        vals = [i, r["db"], r["model"]]
        for mc in metric_cols:
            vals.append(round(r["metrics"].get(mc, 0.0), 4))
        vals.append(round(r["search_time"], 3))
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.alignment = center
            cell.border = border
            if i <= 3:
                cell.fill = gold
        row += 1

    ws.column_dimensions["A"].width = 6
    for c_idx in range(2, 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 18
    for c_idx in range(4, 4 + len(metric_cols) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 13

    # Her DB icin model x metrik tablosu
    for db in DBS:
        ws = wb.create_sheet(f"BB_{db.upper()}")
        ws["A1"] = f"{db.upper()} — Bulma Basarimi (Model x Metrik)"
        ws["A1"].font = Font(bold=True, size=14)

        headers = ["Model"] + metric_cols
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = center
            c.border = border

        row = 4
        for model_key in MODELS:
            ws.cell(row=row, column=1, value=model_key).border = border
            payload = results.get("quality_search", {}).get(db, {}).get(model_key, {})
            metrics = payload.get("metrics", {}) if payload.get("status") == "ok" else {}
            for j, mc in enumerate(metric_cols, 2):
                cell = ws.cell(row=row, column=j)
                cell.value = round(metrics[mc], 4) if mc in metrics else "-"
                cell.alignment = center
                cell.border = border
            row += 1

        ws.column_dimensions["A"].width = 18
        for c_idx in range(2, 2 + len(metric_cols)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 13

    # 6) PERFORMANS METRIKLERI (computed: QPS, p99 proxy, write rec/s, VRAM peak)
    comp = results.get("computed", {})

    # 6a) Write throughput
    ws = wb.create_sheet("Write Throughput")
    ws["A1"] = "YAZMA HIZI (kayit/saniye) + VRAM Peak (GB)"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Model", "DB", "kayit/sn", "sure(s)", "kayit", "vram_peak_gb"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    row = 4
    for model_key in MODELS:
        for db in DBS:
            wd = results.get("write", {}).get(db, {}).get(model_key, {})
            if wd.get("status") != "ok":
                continue
            t = wd.get("time"); rc = wd.get("record_count")
            if not (isinstance(t, (int, float)) and t > 0 and rc):
                continue
            comp_w = comp.get("write", {}).get(db, {}).get(model_key, {})
            vals = [model_key, db,
                    round(rc / t, 1), round(t, 3), rc,
                    round(comp_w.get("vram_peak_gb", 0.0), 2)]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=j, value=v)
                cell.alignment = center; cell.border = border
            row += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    for c in ["C", "D", "E", "F"]:
        ws.column_dimensions[c].width = 14

    # 6b) Search QPS / p99 proxy
    ws = wb.create_sheet("Search QPS p99")
    ws["A1"] = "ARAMA QPS + p99 (max proxy) — TEST_RUNS=10 oldugu icin p99 ≈ max_time"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Sira", "DB", "Model", "Algoritma",
               "QPS", "avg_ms", "p50_ms", "p95_ms", "p99_proxy_ms", "vram_peak_gb"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

    rows_qps = []
    for db in DBS:
        for model_key, algos in comp.get("search", {}).get(db, {}).items():
            if not isinstance(algos, dict):
                continue
            vram = algos.get("__vram_peak_gb", 0.0)
            for algo, m in algos.items():
                if algo.startswith("__") or not isinstance(m, dict):
                    continue
                rows_qps.append({
                    "db": db, "model": model_key, "algo": algo,
                    "qps": m.get("qps", 0.0),
                    "avg": m.get("avg_ms", 0.0),
                    "p50": m.get("p50_ms", 0.0),
                    "p95": m.get("p95_ms", 0.0),
                    "p99": m.get("p99_proxy_ms", 0.0),
                    "vram": vram,
                })
    rows_qps.sort(key=lambda x: x["qps"], reverse=True)

    row = 4
    for i, r in enumerate(rows_qps, 1):
        vals = [i, r["db"], r["model"], r["algo"],
                round(r["qps"], 1), round(r["avg"], 3),
                round(r["p50"], 3), round(r["p95"], 3), round(r["p99"], 3),
                round(r["vram"], 2)]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v)
            cell.alignment = center; cell.border = border
            if i <= 3:
                cell.fill = gold
        row += 1
    ws.column_dimensions["A"].width = 6
    for c in ["B", "C", "D"]:
        ws.column_dimensions[c].width = 18
    for c in ["E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[c].width = 13

    # 6c) Encode VRAM (model bazli, encode + encode_qual)
    ws = wb.create_sheet("Encode VRAM")
    ws["A1"] = "ENCODE FAZI VRAM PEAK (GB) — log HEARTBEAT'ten"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Model", "encode (full corpus)", "encode_qual (SQuAD)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    enc = comp.get("vram_peak_gb", {}).get("encode", {})
    encq = comp.get("vram_peak_gb", {}).get("encode_qual", {})
    row = 4
    for model_key in MODELS:
        vals = [model_key,
                round(enc.get(model_key, 0.0), 2),
                round(encq.get(model_key, 0.0), 2)]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=j, value=v if v else "-")
            cell.alignment = center; cell.border = border
        row += 1
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22

    wb.save(OUTPUT_XLSX)


# ---------------------------------------------------------------
# Entry
# ---------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Vektor DB Multi-Model Benchmark (subprocess izolasyonlu)")
    parser.add_argument("--mode", choices=["test", "full"], default="test")
    parser.add_argument("--reset", action="store_true", help="cache + sonuc dosyasini sifirla")
    parser.add_argument("--worker", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--phase",
                        choices=["encode", "write", "search",
                                 "encode_qual", "write_qual", "search_qual"],
                        help=argparse.SUPPRESS)
    parser.add_argument("--model", help=argparse.SUPPRESS)
    parser.add_argument("--db", choices=DBS, help=argparse.SUPPRESS)
    parser.add_argument("--result-file", dest="result_file", help=argparse.SUPPRESS)
    args = parser.parse_args()

    if args.worker:
        worker_main(args)
        return

    # Orchestrator
    log_file = setup_logging("orchestrator")
    sys.stdout = _PrintToLogger(logging.INFO)
    sys.stderr = _PrintToLogger(logging.ERROR)
    logger.info(f"Log dosyasi: {log_file}")
    logger.info(f"Orchestrator PID={os.getpid()} mode={args.mode} reset={args.reset}")

    inhibit_proc = _inhibit_suspend()
    threading.Thread(target=_heartbeat, daemon=True).start()

    print("=" * 70)
    print(f"  VEKTOR DATABASE BENCHMARK — mod: {args.mode.upper()}")
    print("=" * 70)

    t_start = time.time()
    try:
        orchestrate(args.mode, reset=args.reset)
        logger.info(f"Benchmark TAMAMLANDI — sure: {(time.time()-t_start)/60:.1f} dk")
    except KeyboardInterrupt:
        logger.warning("Kullanici iptali (Ctrl+C) — cikiliyor")
    except Exception as e:
        logger.exception(f"Benchmark FATAL hata: {e}")
        raise
    finally:
        if inhibit_proc is not None:
            inhibit_proc.terminate()


if __name__ == "__main__":
    main()
